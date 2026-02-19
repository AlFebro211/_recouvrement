from urllib import request
from django.http import HttpResponse, JsonResponse
from django.contrib.auth.decorators import login_required
from app.models import *
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Image, Paragraph, Table, TableStyle, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch, mm, cm
from reportlab.lib.pagesizes import A4
from django.conf import settings
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from datetime import datetime, date
import os
from app.models import *
from django.contrib.staticfiles import finders
from django.db.models import Sum
from io import BytesIO
import json
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

def build_pdf_header(eleve=None, classe_obj=None, id_campus=None, id_cycle=None, id_annee=None, titre=None, banque_info=None):
    """
    Header professionnel pour PDF - Avec support de la banque
    """
    
    # Styles professionnels (inchangés)
    style_nom_ecole = ParagraphStyle(
        'NomEcole',
        fontSize=18,
        fontName='Helvetica-Bold',
        alignment=TA_CENTER,
        textColor=colors.HexColor('#1e3c72'),
        leading=22
    )
    
    style_devise = ParagraphStyle(
        'Devise',
        fontSize=11,
        fontName='Helvetica-Oblique',
        alignment=TA_CENTER,
        textColor=colors.HexColor('#666666'),
        leading=14
    )
    
    style_contact = ParagraphStyle(
        'Contact',
        fontSize=9,
        fontName='Helvetica',
        alignment=TA_CENTER,
        textColor=colors.HexColor('#666666'),
        leading=14
    )
    
    style_titre = ParagraphStyle(
        'Titre',
        fontSize=16,
        fontName='Helvetica-Bold',
        alignment=TA_CENTER,
        textColor=colors.HexColor('#2c3e50'),
        spaceAfter=10,
        leading=20,
        backColor=colors.HexColor('#f8fafc'),
        borderPadding=8,
        borderRadius=3
    )
    
    style_info = ParagraphStyle(
        'Info',
        fontSize=10,
        fontName='Helvetica',
        alignment=TA_LEFT,
        textColor=colors.HexColor('#333333'),
        leading=18
    )
    
    style_info_bold = ParagraphStyle(
        'InfoBold',
        fontSize=10,
        fontName='Helvetica-Bold',
        alignment=TA_LEFT,
        textColor=colors.HexColor('#1e3c72'),
        leading=18
    )
    
    style_date = ParagraphStyle(
        'Date',
        fontSize=10,
        fontName='Helvetica',
        alignment=TA_RIGHT,
        textColor=colors.HexColor('#333333'),
        leading=18
    )
    
    style_separator = ParagraphStyle(
        'Separator',
        fontSize=4,
        fontName='Helvetica',
        alignment=TA_CENTER,
        textColor=colors.HexColor('#cccccc')
    )

    # Récupération des informations de l'institution
    institution = None
    try:
        institution = Institution.objects.first()
    except:
        pass
    
    # Initialisation des variables
    nom_ecole = "ECOLE INTERNATIONALE DE BUJUMBURA"
    sigle = ""
    telephone = ""
    email = ""
    site = ""
    siege = ""
    b_postale = ""
    emplacement = ""
    
    if institution:
        if institution.nom_ecole:
            nom_ecole = institution.nom_ecole.upper()
        if institution.sigle:
            sigle = f" - {institution.sigle}"
        if institution.telephone:
            telephone = institution.telephone
        if institution.email:
            email = institution.email
        if institution.site:
            site = institution.site
        if institution.siege:
            siege = institution.siege
        if institution.b_postale:
            b_postale = f"BP {institution.b_postale}"
        if institution.emplacement:
            emplacement = institution.emplacement

    # Construction de l'adresse complète
    adresse_parts = []
    if siege:
        adresse_parts.append(siege)
    if emplacement:
        adresse_parts.append(emplacement)
    if b_postale:
        adresse_parts.append(b_postale)
    
    adresse = " | ".join(filter(None, adresse_parts)) if adresse_parts else "Bujumbura, Burundi"
    
    # Construction des coordonnées complètes
    contact_parts = []
    if telephone:
        contact_parts.append(telephone)
    if email:
        contact_parts.append(email)
    if site:
        contact_parts.append(site)
    
    contact = " | ".join(contact_parts) if contact_parts else ""

    # Logo
    logo_element = Paragraph("", ParagraphStyle('Logo', fontSize=1, alignment=TA_CENTER))
    
    if institution and institution.logo_ecole:
        try:
            logo_path = os.path.join(settings.MEDIA_ROOT, institution.logo_ecole.name)
            if os.path.exists(logo_path):
                logo_element = Image(logo_path, width=2*cm, height=2*cm)
        except:
            pass

    # Informations institution
    institution_text = [
        Paragraph(f"{nom_ecole}{sigle}", style_nom_ecole),
        Spacer(1, 2),
    ]
    
    if adresse:
        institution_text.append(Paragraph(adresse, style_contact))
    if contact:
        institution_text.append(Paragraph(contact, style_contact))

    # Colonne de gauche (Campus et Classe)
    left_info = []
    
    campus_text = "Tous"
    if id_campus:
        try:
            campus = Campus.objects.get(id_campus=id_campus)
            campus_text = campus.campus
        except:
            pass
    left_info.append(Paragraph(f"<b>Campus:</b> {campus_text}", style_info_bold))
    
    classe_text = "Toutes"
    if classe_obj:
        try:
            classe_text = classe_obj.classe_id.classe
            if classe_obj.groupe:
                classe_text += f" ({classe_obj.groupe})"
        except:
            pass
    left_info.append(Paragraph(f"<b>Classe:</b> {classe_text}", style_info_bold))
    
    if eleve:
        nom_complet = f"{eleve.nom} {eleve.prenom}".strip()
        left_info.append(Spacer(1, 3))
        left_info.append(Paragraph(f"<b>Élève:</b> {nom_complet}", style_info_bold))
        if hasattr(eleve, 'matricule') and eleve.matricule:
            left_info.append(Paragraph(f"<b>Matricule:</b> {eleve.matricule}", style_info))

    # Colonne de droite (Année scolaire et Date)
    right_info = []
    
    annee_text = "Non spécifiée"
    if id_annee:
        try:
            annee = Annee.objects.get(id_annee=id_annee)
            annee_text = annee.annee
        except:
            pass
    right_info.append(Paragraph(f"<b>Année scolaire:</b> {annee_text}", style_date))
    right_info.append(Paragraph(f"<b>Date:</b> {datetime.now().strftime('%d/%m/%Y %H:%M')}", style_date))

    # Construction du tableau
    data = []
    
    # LIGNE 1: Logo et informations institution
    data.append([logo_element, institution_text, Paragraph("", style_info)])
    
    # LIGNE 2: Séparateur
    data.append([Paragraph("", style_info), Paragraph("─" * 90, style_separator), Paragraph("", style_info)])
    
    # LIGNE 3: Informations contextuelles (Campus, Classe à gauche / Année, Date à droite)
    data.append([Paragraph("", style_info), left_info, right_info])
    
    # LIGNE 4: Banque (si fournie) - AJOUTÉ ICI, APRÈS LES INFOS CONTEXTUELLES
    if banque_info:
        banque_paragraph = Paragraph(f"<b>Banque sélectionnée :</b> {banque_info}", style_info_bold)
        data.append([Paragraph("", style_info), banque_paragraph, Paragraph("", style_info)])
        # Petit espace après la banque
        data.append([Paragraph("", style_info), Spacer(1, 3), Paragraph("", style_info)])
    
    # LIGNE 5: Espace avant le titre
    data.append([Paragraph("", style_info), Spacer(1, 5), Paragraph("", style_info)])
    
    # LIGNE 6: Titre
    if titre:
        data.append([Paragraph("", style_info), Paragraph(f"<b>{titre.upper()}</b>", style_titre), Paragraph("", style_info)])
    
    # Largeurs des colonnes
    col_widths = [2.5*cm, 12*cm, 4.5*cm]
    
    table = Table(data, colWidths=col_widths)
    
    style_commands = [
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('ALIGN', (0, 0), (0, -1), 'CENTER'),
        ('ALIGN', (1, 0), (1, -1), 'LEFT'),  # Aligné à gauche pour la banque
        ('ALIGN', (2, 0), (2, -1), 'RIGHT'),
        ('LEFTPADDING', (0, 0), (-1, -1), 5),
        ('RIGHTPADDING', (0, 0), (-1, -1), 5),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
    ]
    
    # Déterminer l'index de la ligne de titre
    titre_index = 5  # Par défaut, si pas de banque
    if banque_info:
        titre_index = 6  # Si banque présente, le titre recule d'une ligne
    
    if titre:
        style_commands.append(('BACKGROUND', (1, titre_index), (1, titre_index), colors.HexColor('#f8fafc')))
    
    table.setStyle(TableStyle(style_commands))
    
    return table


def build_pdf_header_pos(eleve=None, classe_obj=None, id_campus=None, id_cycle=None, id_annee=None, titre=None):
    """
    Version POS du header professionnel - Chaque info sur sa propre ligne
    """
    
    # Tailles pour POS
    font_size_normal = 8
    font_size_bold = 10
    font_size_small = 7
    logo_size = 12*mm
    col_widths = [2*cm, 5*cm, 2*cm]
    
    # Styles
    style_nom_ecole = ParagraphStyle(
        'NomEcole',
        fontSize=font_size_bold,
        fontName='Helvetica-Bold',
        alignment=TA_CENTER,
        textColor=colors.HexColor('#1e3c72'),
        leading=12
    )
    
    style_contact = ParagraphStyle(
        'Contact',
        fontSize=font_size_small,
        fontName='Helvetica',
        alignment=TA_CENTER,
        textColor=colors.HexColor('#666666'),
        leading=9
    )
    
    style_titre = ParagraphStyle(
        'Titre',
        fontSize=font_size_bold,
        fontName='Helvetica-Bold',
        alignment=TA_CENTER,
        textColor=colors.HexColor('#2c3e50'),
        leading=12,
        backColor=colors.HexColor('#f8fafc'),
        borderPadding=4,
        borderRadius=2
    )
    
    style_info = ParagraphStyle(
        'Info',
        fontSize=font_size_normal,
        fontName='Helvetica',
        alignment=TA_LEFT,
        textColor=colors.HexColor('#333333'),
        leading=10
    )
    
    style_info_bold = ParagraphStyle(
        'InfoBold',
        fontSize=font_size_normal,
        fontName='Helvetica-Bold',
        alignment=TA_LEFT,
        textColor=colors.HexColor('#1e3c72'),
        leading=10
    )
    
    style_label = ParagraphStyle(
        'Label',
        fontSize=font_size_normal,
        fontName='Helvetica-Bold',
        alignment=TA_LEFT,
        textColor=colors.HexColor('#1e3c72'),
        leading=10
    )
    
    style_valeur = ParagraphStyle(
        'Valeur',
        fontSize=font_size_normal,
        fontName='Helvetica',
        alignment=TA_LEFT,
        textColor=colors.HexColor('#333333'),
        leading=10
    )
    
    style_date = ParagraphStyle(
        'Date',
        fontSize=font_size_small,
        fontName='Helvetica',
        alignment=TA_LEFT,
        textColor=colors.HexColor('#333333'),
        leading=9
    )

    # Récupération institution
    institution = None
    try:
        institution = Institution.objects.first()
    except:
        pass
    
    # Logo
    logo_element = Paragraph("", ParagraphStyle('Logo', fontSize=1))
    if institution and institution.logo_ecole:
        try:
            logo_path = os.path.join(settings.MEDIA_ROOT, institution.logo_ecole.name)
            if os.path.exists(logo_path):
                logo_element = Image(logo_path, width=logo_size, height=logo_size)
        except:
            pass
    
    # Infos institution
    nom_ecole = institution.nom_ecole.upper() if institution and institution.nom_ecole else "ECOLE INTERNATIONALE"
    sigle = f" - {institution.sigle}" if institution and institution.sigle else ""
    
    # Coordonnées
    contact_parts = []
    if institution and institution.telephone:
        contact_parts.append(institution.telephone)
    if institution and institution.email:
        contact_parts.append(institution.email)
    contact = " | ".join(contact_parts) if contact_parts else ""
    
    # Adresse
    adresse_parts = []
    if institution and institution.siege:
        adresse_parts.append(institution.siege)
    if institution and institution.emplacement:
        adresse_parts.append(institution.emplacement)
    if institution and institution.b_postale:
        adresse_parts.append(f"BP {institution.b_postale}")
    adresse = " | ".join(adresse_parts) if adresse_parts else ""
    
    # Construction des données du tableau
    data = []
    
    # LIGNE 1: Logo et nom école
    data.append([
        logo_element,
        Paragraph(f"{nom_ecole}{sigle}", style_nom_ecole),
        Paragraph("", style_info)
    ])
    
    # LIGNE 2: Contact
    if contact:
        data.append([
            Paragraph("", style_info),
            Paragraph(contact, style_contact),
            Paragraph("", style_info)
        ])
    
    # LIGNE 3: Adresse
    if adresse:
        data.append([
            Paragraph("", style_info),
            Paragraph(adresse, style_contact),
            Paragraph("", style_info)
        ])
    
    # LIGNE 4: Séparateur
    data.append([
        Paragraph("", style_info),
        Paragraph("─" * 35, ParagraphStyle('Sep', fontSize=4, alignment=TA_CENTER, textColor=colors.HexColor('#cccccc'))),
        Paragraph("", style_info)
    ])
    
    # LIGNE 5: Campus (sur sa propre ligne)
    if id_campus:
        try:
            campus = Campus.objects.get(id_campus=id_campus)
            data.append([
                Paragraph("", style_info),
                Paragraph(f"<b>Campus:</b> {campus.campus}", style_label),
                Paragraph("", style_info)
            ])
        except:
            pass
    
    # LIGNE 6: Classe (sur sa propre ligne)
    if classe_obj:
        try:
            classe_info = classe_obj.classe_id.classe
            if classe_obj.groupe:
                classe_info += f" ({classe_obj.groupe})"
            data.append([
                Paragraph("", style_info),
                Paragraph(f"<b>Classe:</b> {classe_info}", style_label),
                Paragraph("", style_info)
            ])
        except:
            pass
    
    # LIGNE 7: Année (sur sa propre ligne)
    if id_annee:
        try:
            annee = Annee.objects.get(id_annee=id_annee)
            data.append([
                Paragraph("", style_info),
                Paragraph(f"<b>Année scolaire:</b> {annee.annee}", style_label),
                Paragraph("", style_info)
            ])
        except:
            pass
    
    # LIGNE 8: Date (sur sa propre ligne)
    data.append([
        Paragraph("", style_info),
        Paragraph(f"<b>Date:</b> {datetime.now().strftime('%d/%m/%Y %H:%M')}", style_label),
        Paragraph("", style_info)
    ])
    
    # LIGNE 9: Séparateur avant élève
    data.append([
        Paragraph("", style_info),
        Paragraph("─" * 35, ParagraphStyle('Sep', fontSize=4, alignment=TA_CENTER, textColor=colors.HexColor('#cccccc'))),
        Paragraph("", style_info)
    ])
    
    # LIGNE 10: Élève si présent
    if eleve:
        nom_complet = f"{eleve.nom} {eleve.prenom}".strip()
        data.append([
            Paragraph("", style_info),
            Paragraph(f"👤 <b>{nom_complet}</b>", style_info_bold),
            Paragraph("", style_info)
        ])
        
        # LIGNE 11: Matricule
        if hasattr(eleve, 'matricule') and eleve.matricule:
            data.append([
                Paragraph("", style_info),
                Paragraph(f"🆔 {eleve.matricule}", style_info),
                Paragraph("", style_info)
            ])
        
        # LIGNE 12: Séparateur
        data.append([
            Paragraph("", style_info),
            Paragraph("─" * 35, ParagraphStyle('Sep', fontSize=4, alignment=TA_CENTER, textColor=colors.HexColor('#cccccc'))),
            Paragraph("", style_info)
        ])
    
    # LIGNE 13: Titre (si fourni)
    if titre:
        data.append([
            Paragraph("", style_info),
            Paragraph(f"<b>{titre.upper()}</b>", style_titre),
            Paragraph("", style_info)
        ])
    
    # Création du tableau
    table = Table(data, colWidths=col_widths)
    
    # Style du tableau
    style_commands = [
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('ALIGN', (0, 0), (0, -1), 'CENTER'),
        ('ALIGN', (1, 0), (1, -1), 'LEFT'),  # Aligné à gauche pour meilleure lisibilité
        ('ALIGN', (2, 0), (2, -1), 'RIGHT'),
        ('LEFTPADDING', (0, 0), (-1, -1), 2),
        ('RIGHTPADDING', (0, 0), (-1, -1), 2),
        ('TOPPADDING', (0, 0), (-1, -1), 2),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
    ]
    
    # Fond pour la ligne de titre
    if titre:
        # Trouver l'index de la ligne de titre (dernière ligne)
        style_commands.append(('BACKGROUND', (1, -1), (1, -1), colors.HexColor('#f8fafc')))
    
    table.setStyle(TableStyle(style_commands))
    
    return table

def generate_invoice(request, id_paiement):
    try:
 
        paiement = Paiement.objects.select_related(
            'id_campus', 'id_cycle_actif', 'id_classe_active', 'id_annee', 'id_variable', 'id_eleve'
        ).get(id_paiement=id_paiement)

        institution = Institution.objects.first()  
        if not institution:
            return HttpResponse("Erreur : Institution non trouvée.", status=404)

      
        campus = paiement.id_campus.campus
        cycle = paiement.id_cycle_actif.cycle_id.cycle
        classe = paiement.id_classe_active.classe_id.classe
        groupe = paiement.id_classe_active.groupe or ''
        annee = f"{paiement.id_annee.annee}"
        variable = paiement.id_variable.variable
        montant = f"{paiement.montant} Fbu"
        eleve = f"{paiement.id_eleve.nom} {paiement.id_eleve.prenom}"
        date_generation = datetime.now().strftime("%d/%m/%Y %H:%M")
        # encaisseur = request.user.get_full_name() or request.user.username

     
        page_width = 100 * mm
        page_height = 150 * mm
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="facture_{id_paiement}.pdf"'
        doc = SimpleDocTemplate(response, pagesize=(page_width, page_height), rightMargin=5*mm, leftMargin=5*mm, topMargin=5*mm, bottomMargin=5*mm)


        styles = getSampleStyleSheet()
        normal_style = ParagraphStyle(name='Normal', fontSize=8, leading=10, wordWrap='CJK') 
        table_style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 3),
            ('RIGHTPADDING', (0, 0), (-1, -1), 3),
        ])

       
        elements = []

      
        logo_cell = []
        if institution and institution.logo_ecole:
            logo_ecole_path = os.path.join(settings.MEDIA_ROOT, institution.logo_ecole.name)
            if os.path.exists(logo_ecole_path):
                logo = Image(logo_ecole_path, width=0.8 * inch, height=0.8 * inch)
                logo_cell.append(logo)
            else:
                logo_cell.append(Paragraph("Logo non disponible", normal_style))
        else:
            logo_cell.append(Paragraph("Aucun logo configuré", normal_style))

       
        classe_info = f"{classe} {groupe}".strip() if groupe else classe
        
        info_text = (
            f"Campus: <font color='black'><b>{campus}</b></font><br/>"
            f"Cycle: <font color='black'><b>{cycle}</b></font><br/>"
            f"Classe: <font color='black'><b>{classe_info}</b></font><br/>"
            f"Année scolaire: <font color='black'><b>{annee}</b></font><br/>"
            f"Elève: <font color='black'><b>{eleve}</b></font><br/>"
            f"Date: <font color='black'><b>{date_generation}</b></font>"
        )
        
       
        info_paragraph = Paragraph(info_text, normal_style)

       
        header_table = Table([
            [info_paragraph, logo_cell]
        ], colWidths=[60*mm, 30*mm])
        header_table.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
        ]))
        elements.append(header_table)

       
        variable_length = len(variable)
        if variable_length > 30:  
            variable_col_width = min(65*mm, page_width - 25*mm)  
            montant_col_width = page_width - variable_col_width - 10*mm  
        else:
            variable_col_width = 50*mm
            montant_col_width = 40*mm

        data = [
            ["Variable", "Montant"],
            [Paragraph(variable, normal_style), montant] 
        ]
        table = Table(data, colWidths=[variable_col_width, montant_col_width])
        table.setStyle(table_style)
        elements.append(Spacer(1, 10*mm))
        elements.append(table)

        # encaisseur_text = Paragraph(f"Encaissé par: {encaisseur}", ParagraphStyle(name='Encaisseur', fontSize=8, alignment=2))
        elements.append(Spacer(1, 10*mm))
        # elements.append(encaisseur_text)
        elements.append(Spacer(1, 5*mm))
        elements.append(Paragraph("Signature: ____________________", ParagraphStyle(name='Signature', fontSize=8, alignment=2)))

        doc.build(elements)
        return response

    except Paiement.DoesNotExist:
        return HttpResponse("Erreur : Paiement non trouvé.", status=404)
    except Exception as e:
        return HttpResponse(f"Erreur serveur : {str(e)}", status=500)
    


def build_page_number(canvas, doc):
    """Ajoute la numérotation des pages en bas à droite."""
    page_num = canvas.getPageNumber()
    text = f"Page {page_num}"
    canvas.setFont("Helvetica", 8)
    canvas.drawRightString(doc.width + doc.rightMargin, doc.bottomMargin, text)


def generate_fiche_paie_classe(request):
    try:
        id_campus = request.GET.get('id_campus')
        id_cycle = request.GET.get('id_cycle')
        id_classe_active = request.GET.get('id_classe_active')
        id_annee = request.GET.get('id_annee')

        if not all([id_campus, id_cycle, id_classe_active, id_annee]):
            return HttpResponse("Erreur : Paramètres requis manquants.", status=400)

        inscriptions = Eleve_inscription.objects.filter(
            id_campus_id=id_campus,
            id_classe_cycle_id=id_cycle,
            id_classe_id=id_classe_active,
            id_annee_id=id_annee,
            status=1
        ).select_related('id_eleve').values('id_eleve__id_eleve', 'id_eleve__nom', 'id_eleve__prenom').distinct()

        if not inscriptions.exists():
            return HttpResponse("Erreur : Aucune inscription validée trouvée.", status=404)

        variables = Paiement.objects.filter(
            id_campus_id=id_campus,
            id_cycle_actif_id=id_cycle,
            id_classe_active_id=id_classe_active,
            id_annee_id=id_annee,
            status=True
        ).values('id_variable__variable').distinct()[:4]

        variable_names = [v['id_variable__variable'] for v in variables]
        if not variable_names:
            return HttpResponse("Erreur : Aucun paiement validé trouvé pour cette combinaison.", status=404)

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="fiche_paie_classe_{id_annee}_{id_classe_active}.pdf"'
        doc = SimpleDocTemplate(
            response,
            pagesize=A4,
            rightMargin=10*mm,
            leftMargin=10*mm,
            topMargin=10*mm,
            bottomMargin=15*mm
        )

        styles = getSampleStyleSheet()
        normal_style = ParagraphStyle(name='Normal', fontSize=8, leading=10, wordWrap='CJK')
        title_style = ParagraphStyle(name='Title', fontSize=12, leading=14, spaceAfter=10, alignment=1)
        table_style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('BACKGROUND', (0, 1), (-1, -2), colors.beige),
            ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 3),
            ('RIGHTPADDING', (0, 0), (-1, -1), 3),
            ('SPAN', (0, -1), (1, -1)),  
        ])

        elements = []

        institution = Institution.objects.first()
        logo_cell = []
        if institution and institution.logo_ecole:
            logo_ecole_path = os.path.join(settings.MEDIA_ROOT, institution.logo_ecole.name)
            if os.path.exists(logo_ecole_path):
                logo = Image(logo_ecole_path, width=0.8 * inch, height=0.8 * inch)
                logo_cell.append(logo)
            else:
                logo_cell.append(Paragraph("Logo non disponible", normal_style))
        else:
            logo_cell.append(Paragraph("Aucun logo configuré", normal_style))

        try:
            annee = Annee.objects.get(id_annee=id_annee).annee
            campus = Campus.objects.get(id_campus=id_campus).campus
            cycle = Classe_cycle_actif.objects.get(id_campus=id_campus, id_annee=id_annee, id_cycle_actif=id_cycle, is_active=True).cycle_id.cycle
            classe = Classe_active.objects.get(id_campus=id_campus, id_annee=id_annee, cycle_id=id_cycle, id_classe_active=id_classe_active, is_active=True)
            classe_info = f"{classe.classe_id.classe} {classe.groupe}".strip() if classe.groupe else classe.classe_id.classe
        except Exception as e:
            return HttpResponse(f"Erreur : Données de classe non trouvées.", status=404)

        # info_text = (
        #     f"Campus: <font color='black'><b>{campus}</b></font><br/>"
        #     f"Cycle: <font color='black'><b>{cycle}</b></font><br/>"
        #     f"Classe: <font color='black'><b>{classe_info}</b></font><br/>"
        #     f"Année scolaire: <font color='black'><b>{annee}</b></font><br/>"
        #     f"Date: <font color='black'><b>{datetime.now().strftime('%d/%m/%Y %H:%M')}</b></font>"
        # )
        # info_paragraph = Paragraph(info_text, normal_style)

        # header_table = Table([[info_paragraph, logo_cell]], colWidths=[120*mm, 70*mm])
        # header_table.setStyle(TableStyle([
        #     ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        #     ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
        # ]))
        # elements.append(header_table)
        header_table = build_pdf_header(
            classe_obj=classe,      # Objet Classe_active
            id_campus=id_campus,
            id_cycle=id_cycle,
            id_annee=id_annee
        )
        elements.append(header_table)

        elements.append(Spacer(1, 5*mm))
        elements.append(Paragraph(f"<font color='black'><b>Situation des paiements effectués</b></font>", title_style))
        
        elements.append(Spacer(1, 10*mm))

        data = [
            [Paragraph("N°", ParagraphStyle(name='Header', fontSize=8, fontName='Helvetica-Bold', wordWrap='CJK'))] +
            [Paragraph("Élève", ParagraphStyle(name='Header', fontSize=8, fontName='Helvetica-Bold', wordWrap='CJK'))] +
            [Paragraph(var_name, ParagraphStyle(name='Header', fontSize=8, fontName='Helvetica-Bold', wordWrap='CJK')) for var_name in variable_names]
        ]

        totals = [0] * len(variable_names)

        for idx, inscription in enumerate(inscriptions, start=1):
            eleve = f"{inscription['id_eleve__nom']} {inscription['id_eleve__prenom']}"
            row = [
                Paragraph(str(idx), ParagraphStyle(name='Normal', fontSize=8, leading=10, wordWrap='CJK')), 
                Paragraph(eleve, ParagraphStyle(name='Normal', fontSize=8, leading=10, wordWrap='CJK'))  
            ]
            for i, var_name in enumerate(variable_names):
                paiement = Paiement.objects.filter(
                    id_eleve__id_eleve=inscription['id_eleve__id_eleve'],
                    id_variable__variable=var_name,
                    id_campus_id=id_campus,
                    id_cycle_actif_id=id_cycle,
                    id_classe_active_id=id_classe_active,
                    id_annee_id=id_annee,
                    status=True
                ).first()
                montant = paiement.montant if paiement else 0
                row.append(f"{montant} Fbu" if montant else "-")
                totals[i] += montant if montant else 0
            data.append(row)

        totals_row = [
            Paragraph("Totaux", ParagraphStyle(name='Total', fontSize=8, fontName='Helvetica-Bold')), ""
        ]
        for total in totals:
            totals_row.append(f"{total} Fbu" if total else "-")
        data.append(totals_row)

        total_width = 190 * mm  
        num_col_width = 10 * mm 
        eleve_col_width = 70 * mm 
        variable_col_width = (total_width - num_col_width - eleve_col_width) / max(1, len(variable_names))
        col_widths = [num_col_width, eleve_col_width] + [variable_col_width] * len(variable_names)

        table = Table(data, colWidths=col_widths)
        table.setStyle(table_style)
        elements.append(table)

        doc.build(elements, onFirstPage=build_page_number, onLaterPages=build_page_number)
        return response

    except Exception as e:
        return HttpResponse(f"Erreur serveur : {str(e)}", status=500)


def get_paiements_eleve(request):
    id_eleve = request.GET.get('id_eleve')
    id_campus = request.GET.get('id_campus')
    id_cycle = request.GET.get('id_cycle')
    id_classe = request.GET.get('id_classe_active')
    id_annee = request.GET.get('id_annee')

    paiements = Paiement.objects.filter(
        id_eleve_id=id_eleve,
        id_campus_id=id_campus,
        id_cycle_actif_id=id_cycle,
        id_classe_active_id=id_classe,
        id_annee_id=id_annee,
        status=True
    ).select_related('id_variable')

    data = []
    for p in paiements:
        # statut = "Payé" if p.montant >= p.id_variable.montant else "Non payé"
        data.append({
            "id_paiement": p.id_paiement,
            "variable": p.id_variable.variable,
            "montant_paye": p.montant
        })
    return JsonResponse(data, safe=False)

def generate_fiche_paie_eleve(request):
    try:
        id_eleve = request.GET.get('id_eleve')
        id_campus = request.GET.get('id_campus')
        id_annee = request.GET.get('id_annee')
        id_classe = request.GET.get('id_classe_active')

        eleve = Eleve.objects.get(id_eleve=id_eleve)
        obj_annee = Annee.objects.get(id_annee=id_annee)
        
        paiements_qs = Paiement.objects.filter(
            id_eleve_id=id_eleve, 
            id_annee_id=id_annee, 
            status=True
        ).select_related('id_campus', 'id_classe_active', 'id_variable')

        classe_info = "N/A"
        campus_nom = "N/A"
        if paiements_qs.exists():
            p_ref = paiements_qs[0]
            campus_nom = p_ref.id_campus.campus
            nom_classe = p_ref.id_classe_active.classe_id.classe 
            groupe = p_ref.id_classe_active.groupe or ""
            classe_info = f"{nom_classe} {groupe}".strip()

        def tri_priorite(nom_variable):
            nom = nom_variable.upper()
            if "INSCRIPTION" in nom: return 1
            if "1ER TRIMESTRE" in nom or "TRANCHE 1" in nom: return 2
            if "2EM" in nom or "2EME" in nom or "TRANCHE 2" in nom: return 3
            if "3EM" in nom or "3EME" in nom or "TRANCHE 3" in nom: return 4
            return 99

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename="Facture_{eleve.nom}.pdf"'
        
        doc = SimpleDocTemplate(response, pagesize=A4, leftMargin=15*mm, rightMargin=15*mm, topMargin=10*mm)
        elements = []
        styles = getSampleStyleSheet()

        # --- LOGO ---
        logo_path = finders.find('assets/img/MonEcoleApp-logo.png')
        logo_img = Image(logo_path, width=25*mm, height=25*mm) if logo_path and os.path.exists(logo_path) else ""
        institution = Institution.objects.first()

        # --- HEADER ---
        # header_data = [[
        #     logo_img,
        #     [
        #         Paragraph(f"<b>{institution.nom_ecole if institution else 'N/A'}</b>", fontSize=9, alignment=TA_CENTER ),
        #         Paragraph(f"CAMPUS : {campus_nom}", ParagraphStyle('H3', fontSize=9, alignment=TA_CENTER)),
        #         Paragraph(f"CLASSE : {classe_info}", ParagraphStyle('H3', fontSize=9, alignment=TA_CENTER)),
        #     ],
        #     Paragraph(f"A-A : {obj_annee.annee}", ParagraphStyle('H3', fontSize=10, alignment=TA_RIGHT))
        # ]]
        # header_table = Table(header_data, colWidths=[30*mm, 110*mm, 40*mm])
        # header_table.setStyle(TableStyle([('VALIGN', (0,0), (-1,-1), 'TOP')]))
        # elements.append(header_table)
        # elements.append(Spacer(1, 10*mm))

        # elements.append(Paragraph(f"<b>Nom et Prénom :</b> {eleve.nom} {eleve.prenom}", styles['Normal']))
        # elements.append(Paragraph(f"<b>Matricule :</b> {getattr(eleve, 'matricule', id_eleve)}", styles['Normal']))
        header_table = build_pdf_header(
            eleve=eleve,
            classe_obj=classe,
            id_campus=id_campus,
            id_annee=id_annee,
            titre="FICHE INDIVIDUELLE DE PAIEMENT"
        )
        elements.append(header_table)
        elements.append(Spacer(1, 5*mm))
        # elements.append(Paragraph("<u>SITUATION DES PAIEMENTS</u>", ParagraphStyle('T', fontSize=14, alignment=TA_CENTER)))
        # elements.append(Spacer(1, 5*mm))

        elements.append(Paragraph("<b>I. FRAIS À PAYER</b>", styles['Normal']))
        prix_qs = list(VariablePrix.objects.filter(id_annee_id=id_annee, id_campus_id=id_campus, id_classe_active_id=id_classe))
        prix_qs.sort(key=lambda x: tri_priorite(x.id_variable.variable))

        total_attendu = 0
        data_attendu = [["Motif / Frais", "Montant Attendu"]]
        for p in prix_qs:
            data_attendu.append([p.id_variable.variable, f"{p.prix:,} Fbu".replace(',', ' ')])
            total_attendu += p.prix
        
        t1 = Table(data_attendu, colWidths=[120*mm, 60*mm])
        t1.setStyle(style_tableau_standard())
        elements.append(t1)
        elements.append(Spacer(1, 5*mm))

        elements.append(Paragraph("<b>II. VERSEMENTS EFFECTUÉS</b>", styles['Normal']))
        paiements_liste = list(paiements_qs)
        paiements_liste.sort(key=lambda x: tri_priorite(x.id_variable.variable))

        total_paye = 0
        data_paye = [["Motif", "Montant Payé", "Date"]]
        for p in paiements_liste:
            data_paye.append([p.id_variable.variable, f"{p.montant:,} Fbu".replace(',', ' '), p.date_paie.strftime('%d/%m/%Y')])
            total_paye += p.montant
        
        t2 = Table(data_paye, colWidths=[80*mm, 50*mm, 50*mm])
        t2.setStyle(style_tableau_standard())
        elements.append(t2)
        elements.append(Spacer(1, 10*mm))

        reste = total_attendu - total_paye
        status = "EN ORDRE" if reste <= 0 else "EN DETTE"
        couleur = colors.green if reste <= 0 else colors.red

        data_recap = [
            ["TOTAL ATTENDU", "TOTAL PAYÉ", "SOLDE / RESTE", "STATUT"],
            [f"{total_attendu:,} Fbu".replace(',', ' '), 
             f"{total_paye:,} Fbu".replace(',', ' '), 
             f"{max(0, reste):,} Fbu".replace(',', ' '), 
             status]
        ]

        t_recap = Table(data_recap, colWidths=[45*mm, 45*mm, 45*mm, 45*mm])
        t_recap.setStyle(TableStyle([
            ('GRID', (0,0), (-1,-1), 1, colors.black),
            ('BACKGROUND', (0,0), (-1,0), colors.lightgrey),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('FONTNAME', (0,0), (-1,-1), 'Helvetica-Bold'),
            ('TEXTCOLOR', (3,1), (3,1), couleur),
        ]))
        elements.append(t_recap)

        doc.build(elements)
        return response

    except Exception as e:
        return HttpResponse(f"Erreur technique : {str(e)}", status=500)

def style_tableau_standard():
    return TableStyle([
        ('GRID', (0,0), (-1,-1), 0.5, colors.black),
        ('BACKGROUND', (0,0), (-1,0), colors.whitesmoke),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('ALIGN', (1,0), (1,-1), 'RIGHT'), # Montants à droite
        ('FONTSIZE', (0,0), (-1,-1), 9),
    ])


def generate_facture_paiement(request):
    try:
        id_paiement = request.GET.get('id_paiement')

        # Récupération paiement
        paiement = Paiement.objects.select_related(
            'id_eleve', 'id_variable', 'id_campus',
            'id_classe_active', 'id_classe_active__classe_id',
            'id_annee', 'id_cycle_actif'
        ).get(id_paiement=id_paiement)

        eleve = paiement.id_eleve
        campus_nom = paiement.id_campus.campus if paiement.id_campus else ""
        nom_classe = paiement.id_classe_active.classe_id.classe if paiement.id_classe_active else ""
        groupe = paiement.id_classe_active.groupe or ""
        classe_info = f"{nom_classe} {groupe}".strip()
        annee_txt = paiement.id_annee.annee if paiement.id_annee else ""

        # =========================
        # Mode POS ou A4
        # =========================
        is_pos = request.GET.get("pos") == "1"

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename="Facture_{eleve.nom}_{eleve.prenom}.pdf"'

        if is_pos:
            pagesize = (80*mm, 350*mm)
            margin = 5*mm
            font_title = 11
            font_text = 9
            font_small = 8
        else:
            pagesize = A4
            margin = 15*mm
            font_title = 16
            font_text = 11
            font_small = 9

        doc = SimpleDocTemplate(
            response,
            pagesize=pagesize,
            leftMargin=margin,
            rightMargin=margin,
            topMargin=8*mm,
            bottomMargin=8*mm
        )

        elements = []

        # Styles
        style_titre = ParagraphStyle("titre", alignment=TA_CENTER, fontSize=font_title+2, fontName='Helvetica-Bold', textColor=colors.HexColor('#2c3e50'), spaceAfter=6, leading=font_title+6, backColor=colors.HexColor('#f8fafc'), borderPadding=6, borderRadius=3)
        style_sous_titre = ParagraphStyle("sous_titre", alignment=TA_LEFT, fontSize=font_text+1, fontName='Helvetica-Bold', textColor=colors.HexColor('#1e3c72'), spaceAfter=4, leading=font_text+4)
        style_label = ParagraphStyle("label", alignment=TA_LEFT, fontSize=font_text, fontName='Helvetica-Bold', textColor=colors.HexColor('#333333'), leading=font_text+3)
        style_valeur = ParagraphStyle("valeur", alignment=TA_LEFT, fontSize=font_text, fontName='Helvetica', textColor=colors.HexColor('#333333'), leading=font_text+3)
        style_footer = ParagraphStyle("footer", alignment=TA_CENTER, fontSize=font_small-1, fontName='Helvetica', textColor=colors.gray, leading=font_small)
        style_separator = ParagraphStyle("separator", alignment=TA_CENTER, fontSize=4, fontName='Helvetica', textColor=colors.HexColor('#cccccc'))

        # =========================
        # HEADER selon le mode
        # =========================
        if is_pos:
            header_pos = build_pdf_header_pos(
                classe_obj=paiement.id_classe_active,
                id_campus=paiement.id_campus.id_campus if paiement.id_campus else None,
                id_cycle=paiement.id_cycle_actif.id_cycle_actif if paiement.id_cycle_actif else None,
                id_annee=paiement.id_annee.id_annee if paiement.id_annee else None,
                # titre="REÇU DE PAIEMENT"
            )
            elements.append(header_pos)
        else:
            # Mode A4 : utiliser le header complet
            header_complet = build_pdf_header(
                classe_obj=paiement.id_classe_active,
                id_campus=paiement.id_campus.id_campus if paiement.id_campus else None,
                id_cycle=paiement.id_cycle_actif.id_cycle_actif if paiement.id_cycle_actif else None,
                id_annee=paiement.id_annee.id_annee if paiement.id_annee else None,
                titre="REÇU DE PAIEMENT"
            )
            elements.append(header_complet)

        elements.append(Spacer(1, 5*mm))

        if is_pos:
            elements.append(Paragraph("REÇU DE PAIEMENT", style_titre))
            elements.append(Spacer(1, 5*mm))

        # Informations élève (UNE SEULE FOIS)
        elements.append(Paragraph("INFORMATIONS ÉLÈVE", style_sous_titre))
        elements.append(Paragraph("─" * 40 if not is_pos else "─" * 30, style_separator))
        elements.append(Spacer(1, 2*mm))

        eleve_data = [
            [Paragraph("<b>Nom & Prénom :</b>", style_label), Paragraph(f"{eleve.nom} {eleve.prenom}", style_valeur)],
            [Paragraph("<b>Matricule :</b>", style_label), Paragraph(eleve.matricule if hasattr(eleve, 'matricule') and eleve.matricule else "N/A", style_valeur)],
            [Paragraph("<b>Classe :</b>", style_label), Paragraph(classe_info if classe_info else "N/A", style_valeur)],
        ]

        eleve_table = Table(eleve_data, colWidths=['30%', '70%'] if not is_pos else [25*mm, 45*mm])
        eleve_table.setStyle(TableStyle([
            ('FONTSIZE', (0,0), (-1,-1), font_text),
            ('BOTTOMPADDING', (0,0), (-1,-1), 3),
            ('TOPPADDING', (0,0), (-1,-1), 3),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ]))
        elements.append(eleve_table)
        elements.append(Spacer(1, 8*mm))

        # Détail du paiement
        elements.append(Paragraph("DÉTAIL DU PAIEMENT", style_sous_titre))
        elements.append(Paragraph("─" * 40 if not is_pos else "─" * 30, style_separator))
        elements.append(Spacer(1, 4*mm))

        montant_formate = f"{paiement.montant:,}".replace(',', ' ') + " FBu"
        
        # Tableau de détail
        detail_data = [
            [Paragraph("<b>Désignation</b>", style_label), Paragraph("<b>Montant</b>", style_label)],
            [paiement.id_variable.variable, montant_formate]
        ]
        
        if is_pos:
            col_widths = [45*mm, 25*mm]
        else:
            col_widths = [120*mm, 60*mm]
            
        detail_table = Table(detail_data, colWidths=col_widths)
        detail_table.setStyle(TableStyle([
            ('GRID', (0,0), (-1,-2), 0.5, colors.black),
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#1e3c72")),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('ALIGN', (0,0), (-1,0), 'CENTER'),
            ('ALIGN', (1,1), (1,1), 'RIGHT'),
            ('FONTSIZE', (0,0), (-1,-1), font_text),
            ('BOTTOMPADDING', (0,0), (-1,-1), 5),
            ('TOPPADDING', (0,0), (-1,-1), 5),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ]))
        elements.append(detail_table)
        elements.append(Spacer(1, 2*mm))

        # TOTAL
        total_table = Table([["TOTAL PAYE", montant_formate]], colWidths=col_widths)
        total_table.setStyle(TableStyle([
            ('LINEABOVE', (0,0), (-1,0), 1, colors.HexColor("#1e3c72")),
            ('ALIGN', (0,0), (0,0), 'LEFT'),
            ('ALIGN', (1,0), (1,0), 'RIGHT'),
            ('FONTSIZE', (0,0), (-1,-1), font_text+2),
            ('FONTNAME', (0,0), (-1,-1), 'Helvetica-Bold'),
            ('BACKGROUND', (0,0), (-1,-1), colors.HexColor("#f8fafc")),
            ('TEXTCOLOR', (1,0), (1,0), colors.HexColor("#1e3c72")),
            ('BOTTOMPADDING', (0,0), (-1,-1), 6),
            ('TOPPADDING', (0,0), (-1,-1), 6),
        ]))
        elements.append(total_table)
        elements.append(Spacer(1, 8*mm))

        # Informations complémentaires
        info_comp_data = [
            [Paragraph("<b>Mode de paiement :</b>", style_label), Paragraph("Espèces", style_valeur)],
            [Paragraph("<b>Date de paiement :</b>", style_label), Paragraph(paiement.date_paie.strftime('%d/%m/%Y'), style_valeur)],
        ]
        
        info_comp_table = Table(info_comp_data, colWidths=['30%', '70%'] if not is_pos else [25*mm, 45*mm])
        info_comp_table.setStyle(TableStyle([
            ('FONTSIZE', (0,0), (-1,-1), font_text),
            ('BOTTOMPADDING', (0,0), (-1,-1), 3),
            ('TOPPADDING', (0,0), (-1,-1), 3),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ]))
        elements.append(info_comp_table)
        elements.append(Spacer(1, 8*mm))

        
        # Date d'édition
        elements.append(Paragraph(
            f"Édité le {datetime.now().strftime('%d/%m/%Y à %H:%M')}",
            ParagraphStyle("date_edition", alignment=TA_CENTER, fontSize=font_small-1, textColor=colors.gray)
        ))

        doc.build(elements)
        return response

    except Paiement.DoesNotExist:
        return HttpResponse("Paiement introuvable", status=404)
    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return HttpResponse(f"Erreur technique : {str(e)}", status=500)

# logos/ecole/eibu_yxQIjf2.PNG

def generate_historique_pdf(request):
    try:
        id_annee = request.GET.get('annee')
        id_classe = request.GET.get('classe')
        id_eleve = request.GET.get('eleve')
        id_trimestre = request.GET.get('trimestre')
        id_compte = request.GET.get('compte')

        # ============================================
        # CONSTRUCTION DE LA REQUÊTE DE BASE
        # ============================================
        paiements_qs = Paiement.objects.select_related(
            'id_eleve',
            'id_variable',
            'id_compte',
            'id_compte__id_banque',
            'id_classe_active',
            'id_classe_active__classe_id',
            'id_classe_active__id_campus',
            'id_annee'
        ).order_by(
            'id_eleve__nom',
            'id_variable__variable',
            'date_paie'
        )

        # Filtres de base
        if id_annee:
            paiements_qs = paiements_qs.filter(id_annee=id_annee)
        if id_classe:
            paiements_qs = paiements_qs.filter(id_classe_active=id_classe)
        if id_eleve:
            paiements_qs = paiements_qs.filter(id_eleve=id_eleve)
        if id_compte:
            paiements_qs = paiements_qs.filter(id_compte=id_compte)

        # ============================================
        # FILTRE PAR TRIMESTRE VIA VARIABLEPRIX
        # ============================================
        if id_trimestre:
            from app.models import VariablePrix
            
            variable_ids = VariablePrix.objects.filter(
                id_annee_trimestre_id=id_trimestre
            ).values_list('id_variable_id', flat=True).distinct()
            
            paiements_qs = paiements_qs.filter(id_variable_id__in=variable_ids)

        # ============================================
        # CRÉATION DU PDF
        # ============================================
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'inline; filename="Rapport_Financier.pdf"'

        doc = SimpleDocTemplate(
            response,
            pagesize=A4,
            leftMargin=10*mm,
            rightMargin=10*mm,
            topMargin=10*mm,
            bottomMargin=10*mm
        )

        elements = []
        styles = getSampleStyleSheet()

        small = ParagraphStyle('small', parent=styles['Normal'], fontSize=8, leading=10)
        left_style = ParagraphStyle('left', fontSize=9, leading=11)

        # Logo
        logo_path = finders.find('assets/img/logo.png')
        if logo_path:
            elements.append(Image(logo_path, width=25*mm, height=25*mm))
        elements.append(Spacer(1, 4*mm))

        # Récupérer la première ligne pour les infos d'en-tête
        p_ref = paiements_qs.first()
        
        # ============================================
        # PRÉPARER LES INFORMATIONS POUR LE HEADER
        # ============================================
        eleve_obj = None
        classe_obj = None
        id_campus_val = None
        id_cycle_val = None
        banque_info = None
        
        if p_ref:
            if p_ref.id_classe_active:
                classe_obj = p_ref.id_classe_active
                if p_ref.id_classe_active.id_campus:
                    id_campus_val = p_ref.id_classe_active.id_campus.id_campus
                if hasattr(p_ref.id_classe_active, 'id_cycle_actif') and p_ref.id_classe_active.id_cycle_actif:
                    id_cycle_val = p_ref.id_classe_active.id_cycle_actif.id_cycle_actif
            
            if id_eleve and p_ref.id_eleve:
                eleve_obj = p_ref.id_eleve
            
            if id_compte and p_ref.id_compte:
                banque_info = f"{p_ref.id_compte.id_banque.banque} - {p_ref.id_compte.compte}"

        # ============================================
        # HEADER AVEC build_pdf_header (MODIFIÉ POUR ACCEPTER BANQUE)
        # ============================================
        header_table = build_pdf_header(
            eleve=eleve_obj,
            classe_obj=classe_obj,
            id_campus=id_campus_val,
            id_cycle=id_cycle_val,
            id_annee=int(id_annee) if id_annee else None,
            titre="HISTORIQUE DES PAIEMENTS",
            banque_info=banque_info  # NOUVEAU paramètre pour la banque
        )
        elements.append(header_table)
        elements.append(Spacer(1, 4*mm))

        # ============================================
        # DÉTERMINER LES COLONNES DU TABLEAU
        # ============================================
        # Si une banque est sélectionnée, on NE l'affiche PAS dans le tableau
        afficher_banque_dans_tableau = not id_compte

        # Tableau
        table_header = ["Élève", "Variable", "Montant", "Date paiement"]
        if afficher_banque_dans_tableau:
            table_header.append("Banque / Compte")
        data = [table_header]

        total_general = 0

        for p in paiements_qs:
            total_general += p.montant
            row = [
                Paragraph(f"{p.id_eleve.nom} {p.id_eleve.prenom}", small),
                Paragraph(p.id_variable.variable, small),
                f"{p.montant:,.0f}",
                Paragraph(p.date_paie.strftime('%d/%m/%Y'), small)
            ]
            if afficher_banque_dans_tableau:
                row.append(Paragraph(f"{p.id_compte.id_banque.banque} - {p.id_compte.compte}" if p.id_compte else "-", small))
            data.append(row)

        # Total général
        total_row = ["", "TOTAL GÉNÉRAL", f"{total_general:,.0f}", ""]
        if afficher_banque_dans_tableau:
            total_row.append("")
        data.append(total_row)

        # Largeurs des colonnes
        colWidths = [45*mm, 40*mm, 25*mm, 30*mm]
        if afficher_banque_dans_tableau:
            colWidths.append(45*mm)

        table = Table(data, colWidths=colWidths, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#0d6efd')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('ALIGN', (2,1), (2,-1), 'RIGHT'),
            ('FONTNAME', (0,-1), (-1,-1), 'Helvetica-Bold'),
            ('BACKGROUND', (0,-1), (-1,-1), colors.HexColor('#ffc107')),
            ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
            ('FONTSIZE', (0,1), (-1,-2), 8),
        ]))

        elements.append(table)
        doc.build(elements)
        return response

    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return HttpResponse(f"Erreur PDF : {str(e)}", status=500)

def generate_dette_pdf(request):
    try:
        id_annee = request.GET.get('annee')
        id_classe = request.GET.get('classe')
        id_campus = request.GET.get('id_campus')
        id_trimestre = request.GET.get('trimestre')

        if not id_annee or not id_classe:
            return HttpResponse("Année et classe obligatoires", status=400)

        # 🔹 1. Récupérer les inscriptions
        inscriptions = Eleve_inscription.objects.select_related(
            'id_eleve', 'id_classe', 'id_classe__classe_id'
        ).filter(
            id_annee=id_annee,
            id_classe=id_classe,
            status=True
        )
        if id_trimestre:
            inscriptions = inscriptions.filter(id_trimestre=id_trimestre)

        rapport = []

        # 🔹 Préparer infos globales pour entête
        classe_obj = Classe_active.objects.filter(classe_id=id_classe).first()
        nom_classe = classe_obj.classe_id.classe if classe_obj else "N/A"
        groupe = classe_obj.groupe if classe_obj and hasattr(classe_obj, 'groupe') else ""
        classe_info = f"{nom_classe} {groupe}".strip()
        campus_txt = classe_obj.id_campus.campus if classe_obj and hasattr(classe_obj, 'id_campus') and classe_obj.id_campus else "N/A"

        annee_obj = Annee.objects.filter(id_annee=id_annee).first()
        annee_txt = annee_obj.annee if annee_obj else "N/A"
        trimestre_txt = id_trimestre or "Tous"

        # 🔹 Construire le rapport
        for ins in inscriptions:
            eleve = ins.id_eleve
            paiements = Paiement.objects.filter(
                id_eleve=eleve,
                id_classe_active=id_classe,
                id_annee=id_annee
            )

            variables = VariablePrix.objects.filter(
                id_classe_active=id_classe,
                id_annee=id_annee
            ).select_related('id_variable')

            details = []
            total_dette = 0

            for vp in variables:
                variable = vp.id_variable
                montant_a_payer = vp.prix
                montant_paye = paiements.filter(id_variable=variable).aggregate(total=Sum('montant'))['total'] or 0
                reste = max(montant_a_payer - montant_paye, 0)

                penalite = 0
                if reste > 0:
                    penalite = PenaliteConfig.objects.filter(
                        id_variable=variable,
                        id_classe_active=id_classe,
                        actif=True
                    ).aggregate(total=Sum('valeur'))['total'] or 0

                total_variable = reste + penalite
                if total_variable > 0:
                    details.append({
                        "variable": variable.variable,
                        "montant_a_payer": montant_a_payer,
                        "montant_paye": montant_paye,
                        "reste": reste,
                        "penalite": penalite,
                        "total": total_variable
                    })
                    total_dette += total_variable

            if total_dette > 0:
                rapport.append({
                    "eleve": f"{eleve.nom} {eleve.prenom}",
                    "classe": classe_info,
                    "total_dette": total_dette,
                    "details": details
                })

        # 🔹 Générer le PDF
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'inline; filename="Dette_Rapport.pdf"'

        doc = SimpleDocTemplate(
            response,
            pagesize=A4,
            leftMargin=10*mm, rightMargin=10*mm,
            topMargin=10*mm, bottomMargin=10*mm
        )

        elements = []
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle('title', fontSize=14, alignment=TA_CENTER, fontName='Helvetica-Bold')
        small_left = ParagraphStyle('small_left', alignment=TA_LEFT, fontSize=9)
        small_right = ParagraphStyle('small_right', alignment=TA_RIGHT, fontSize=9)

        # Logo
        # 🔹 Logo
        logo_path = finders.find('assets/img/logo.png')
        logo_img = Image(logo_path, width=30*mm, height=30*mm) if logo_path and os.path.exists(logo_path) else Paragraph("", styles['Normal'])
        institution = Institution.objects.first()
        # 🔹 Entête améliorée
        # header_data = [[
        #     logo_img,
        #     [
        #         Paragraph(f"<b>{institution.nom_ecole if institution else 'N/A'}</b>", title_style),
        #         Paragraph(f"CAMPUS : {campus_txt}", ParagraphStyle('H3', fontSize=9, alignment=TA_LEFT)),
        #         Paragraph(f"CLASSE : {classe_info}", ParagraphStyle('H3', fontSize=9, alignment=TA_LEFT)),
        #         Paragraph(f"Trimestre : {id_trimestre or 'Tous'}", ParagraphStyle('H3', fontSize=9, alignment=TA_LEFT)),
        #     ],
        #     Paragraph(f"A-A : {annee_txt}", ParagraphStyle('H3', fontSize=10, alignment=TA_RIGHT))
        # ]]
        header_table = build_pdf_header(
            classe_obj=classe_obj,
            id_campus=classe_obj.id_campus.id_campus,
            id_annee=id_annee,
            titre="RAPPORT DES DETTES"
        )
        elements.append(header_table)
        elements.append(Spacer(1, 5*mm))
        # header_table = Table(header_data, colWidths=[30*mm, 110*mm, 40*mm])
        header_table.setStyle(TableStyle([('VALIGN', (0,0), (-1,-1), 'TOP')]))
        # elements.append(header_table)
        elements.append(Spacer(1, 10*mm))


        # Titre
        # elements.append(Paragraph("RAPPORT DES DETTES", title_style))
        # elements.append(Spacer(1,3*mm))

        # Tableau
        data = [["Élève","Variable","À payer","Payé","Reste","Pénalité","Total"]]
        for eleve in rapport:
            details = eleve['details']
            for i, d in enumerate(details):
                row = []
                if i == 0:
                    row.append(Paragraph(f"<b>{eleve['eleve']}</b><br/><font color='red'>Dette: {eleve['total_dette']:,} FBU</font>", styles['Normal']))
                else:
                    row.append('')
                row.append(d['variable'])
                row.append(f"{d['montant_a_payer']:,}")
                row.append(f"{d['montant_paye']:,}")
                row.append(f"{d['reste']:,}")
                row.append(f"{d['penalite']:,}")
                row.append(f"{d['total']:,}")
                data.append(row)

        table = Table(data, colWidths=[40*mm,35*mm,20*mm,20*mm,20*mm,20*mm,25*mm], repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND',(0,0),(-1,0),colors.HexColor('#0d6efd')),
            ('TEXTCOLOR',(0,0),(-1,0),colors.white),
            ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),
            ('ALIGN',(2,1),(-1,-1),'RIGHT'),
            ('GRID',(0,0),(-1,-1),0.5,colors.grey),
            ('FONTSIZE',(0,1),(-1,-1),8),
            ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
        ]))
        elements.append(table)
        elements.append(Spacer(1,6*mm))

        doc.build(elements)
        return response

    except Exception as e:
        return HttpResponse(f"Erreur PDF : {str(e)}", status=500)

def generate_situation_pdf(request):
    try:
        date_debut = request.GET.get('date_debut')
        date_fin = request.GET.get('date_fin')

        if not date_debut or not date_fin:
            return HttpResponse("Erreur : veuillez fournir les deux dates.", status=400)

        paiements_qs = Paiement.objects.select_related(
            'id_eleve','id_variable','id_banque','id_compte',
            'id_campus', 'id_annee', 'id_classe_active', 'id_cycle_actif'
        ).filter(date_saisie__range=[date_debut, date_fin]).order_by(
            'id_eleve__nom','id_variable__variable','date_saisie'
        )

        # ======================
        # Récupérer les informations pour le header
        # ======================
        premier_paiement = paiements_qs.first()
        campus_id = None
        annee_id = None
        
        if premier_paiement:
            campus_id = premier_paiement.id_campus.id_campus if premier_paiement.id_campus else None
            annee_id = premier_paiement.id_annee.id_annee if premier_paiement.id_annee else None

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'inline; filename="Situation_Journaliere.pdf"'

        doc = SimpleDocTemplate(response, pagesize=A4, leftMargin=10*mm, rightMargin=10*mm, topMargin=10*mm, bottomMargin=10*mm)
        elements = []
        styles = getSampleStyleSheet()

        # ======================
        # HEADER AVEC build_pdf_header
        # ======================
        header = build_pdf_header(
            id_campus=campus_id,
            id_annee=annee_id,
            titre="SITUATION JOURNALIÈRE DES PAIEMENTS"
        )
        elements.append(header)
        elements.append(Spacer(1, 8*mm))

        # ======================
        # Informations de période
        # ======================
        periode_style = ParagraphStyle(
            'Periode',
            parent=styles['Normal'],
            fontSize=10,
            alignment=TA_CENTER,
            textColor=colors.HexColor('#666666'),
            spaceAfter=6
        )
        elements.append(Paragraph(f"Période du {date_debut} au {date_fin}", periode_style))
        elements.append(Spacer(1, 5*mm))

        # ======================
        # TABLEAU
        # ======================
        small = ParagraphStyle('small', parent=styles['Normal'], fontSize=8, leading=10)
        
        data = [["Élève", "Variable", "Montant", "Date paiement", "Compte"]]
        total_general = 0

        for p in paiements_qs:
            total_general += p.montant
            data.append([
                Paragraph(f"{p.id_eleve.nom} {p.id_eleve.prenom}", small),
                Paragraph(p.id_variable.variable, small),
                f"{p.montant:,.0f}",
                Paragraph(p.date_paie.strftime('%d/%m/%Y'), small),
                Paragraph(f"{p.id_compte.compte} - {p.id_banque.banque}", small)
            ])

        # Total général
        data.append(["", "TOTAL", f"{total_general:,.0f}", "", ""])

        table = Table(data, colWidths=[50*mm, 40*mm, 30*mm, 30*mm, 40*mm])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#0d6efd')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('ALIGN', (2,1), (2,-1), 'RIGHT'),
            ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('BACKGROUND', (0,-1), (-1,-1), colors.HexColor('#ffc107')),
        ]))

        elements.append(table)
        
        # ======================
        # Pied de page avec totaux
        # ======================
        elements.append(Spacer(1, 5*mm))
        footer_style = ParagraphStyle(
            'Footer',
            parent=styles['Normal'],
            fontSize=8,
            alignment=TA_RIGHT,
            textColor=colors.HexColor('#666666')
        )
        elements.append(Paragraph(f"Total général: {total_general:,.0f} FBU", footer_style))

        doc.build(elements)
        return response

    except Exception as e:
        return HttpResponse(f"Erreur PDF : {str(e)}", status=500)


def export_dashboard_pdf(request):
    data = json.loads(request.body)

    title = data.get("title", "Tableau")
    headers = data.get("headers", [])
    rows = data.get("rows", [])
    total_general = data.get("total", 0)

    id_campus = data.get("id_campus")
    id_cycle = data.get("id_cycle")
    id_annee = data.get("id_annee")
    id_classe = data.get("id_classe")
    id_eleve = data.get("id_eleve")

    # Récupération des objets
    classe_obj = None
    eleve_obj = None

    if id_classe:
        classe_obj = Classe_active.objects.filter(id_classe_active=id_classe).first()

    if id_eleve:
        eleve_obj = Eleve.objects.filter(id_eleve=id_eleve).first()

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)

    elements = []

    # Header
    header = build_pdf_header(
        eleve=eleve_obj,
        classe_obj=classe_obj,
        id_campus=id_campus,
        id_cycle=id_cycle,
        id_annee=id_annee,
        titre=title
    )
    elements.append(header)
    elements.append(Spacer(1, 15))

    styles = getSampleStyleSheet()
    
    # Nettoyer les rows
    cleaned_rows = []
    for row in rows:
        cleaned_row = []
        for cell in row:
            cell = str(cell).strip()
            cleaned_row.append(cell)
        cleaned_rows.append(cleaned_row)

    # Construction du tableau
    nb_colonnes = len(headers)
    
    # Formater le total avec des espaces
    total_formate = f"{int(total_general):,}".replace(',', ' ')
    
    # Créer la ligne de total avec le libellé "TOTAL GÉNÉRAL"
    if nb_colonnes == 4:  # Cas avec Élève
        total_row = ['', '', 'TOTAL GÉNÉRAL', total_formate]
    elif nb_colonnes == 3:  # Cas sans Élève
        total_row = ['', 'TOTAL GÉNÉRAL', total_formate]
    elif nb_colonnes == 2:
        total_row = ['TOTAL GÉNÉRAL', total_formate]
    else:
        total_row = [''] * (nb_colonnes - 2) + ['TOTAL GÉNÉRAL', total_formate]
    
    # Ajouter la ligne de total
    table_data = [headers] + cleaned_rows + [total_row]

    # Créer le tableau
    table = Table(table_data, repeatRows=1)

    # Style du tableau
    style = TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.grey),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('ALIGN', (0,0), (-1,0), 'CENTER'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,0), 10),
        ('BOTTOMPADDING', (0,0), (-1,0), 12),
        ('BACKGROUND', (0,1), (-1,-2), colors.white),
        ('GRID', (0,0), (-1,-2), 1, colors.black),
        ('ALIGN', (-1,1), (-1,-2), 'RIGHT'),
        ('BACKGROUND', (0,-1), (-1,-1), colors.lightgrey),
        ('FONTNAME', (0,-1), (-1,-1), 'Helvetica-Bold'),
        ('GRID', (0,-1), (-1,-1), 1, colors.black),
        ('ALIGN', (-1,-1), (-1,-1), 'RIGHT'),
    ])
    
    # Fusionner les cellules de la ligne de total
    if nb_colonnes > 2:
        style.add('SPAN', (0, -1), (nb_colonnes-2, -1))
    
    table.setStyle(style)

    elements.append(table)

    doc.build(elements)

    pdf = buffer.getvalue()
    buffer.close()

    filename = f"{title.replace(' ', '_').lower()}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"

    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    return response

def generate_historique_excel(request):
    try:
        annee = request.GET.get('annee')
        classe = request.GET.get('classe')
        trimestre = request.GET.get('trimestre')
        eleve = request.GET.get('eleve')
        compte = request.GET.get('compte')

        # ============================================
        # CONSTRUCTION DE LA REQUÊTE DE BASE
        # ============================================
        paiements = Paiement.objects.select_related(
            'id_eleve',
            'id_variable',
            'id_compte',
            'id_compte__id_banque'
        ).filter(
            id_annee=annee,
            id_classe_active=classe
        ).order_by('id_eleve__nom', 'id_eleve__prenom')

        if eleve:
            paiements = paiements.filter(id_eleve=eleve)

        if compte:
            paiements = paiements.filter(id_compte=compte)

        # ============================================
        # FILTRE PAR TRIMESTRE VIA VARIABLEPRIX
        # ============================================
        if trimestre:
            from app.models import VariablePrix
            
            variable_ids = VariablePrix.objects.filter(
                id_annee_trimestre_id=trimestre
            ).values_list('id_variable_id', flat=True).distinct()
            
            paiements = paiements.filter(id_variable_id__in=variable_ids)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Rapport Financier"

        # Titre
        ws.merge_cells('A1:E1')
        ws['A1'] = "RAPPORT FINANCIER - HISTORIQUE DES PAIEMENTS"
        ws['A1'].font = Font(size=14, bold=True)
        ws['A1'].alignment = Alignment(horizontal="center")

        # Informations de filtre (optionnel)
        ws.merge_cells('A2:E2')
        filter_info = []
        if annee:
            filter_info.append(f"Année: {annee}")
        if classe:
            filter_info.append(f"Classe: {classe}")
        if trimestre:
            filter_info.append(f"Trimestre: {trimestre}")
        if eleve:
            filter_info.append(f"Élève: {eleve}")
        if compte:
            filter_info.append(f"Compte: {compte}")
        
        ws['A2'] = " | ".join(filter_info) if filter_info else "Tous les filtres"
        ws['A2'].font = Font(italic=True)
        ws['A2'].alignment = Alignment(horizontal="center")

        # Ligne vide
        ws.append([])

        # Entêtes
        headers = ["Élève", "Variable", "Montant payé", "Date paiement", "Banque / Compte"]
        ws.append(headers)

        for cell in ws[4]:  # Ligne 4 car les 3 premières sont occupées
            cell.font = Font(bold=True)

        total_general = 0

        # Données
        for p in paiements:
            montant = p.montant or 0
            total_general += montant

            ws.append([
                str(p.id_eleve),
                str(p.id_variable.variable),
                montant,
                p.date_paie.strftime("%d/%m/%Y") if p.date_paie else "",
                f"{p.id_compte.id_banque.banque} - {p.id_compte.compte}" if p.id_compte else "-"
            ])

        # Ligne total
        ws.append(["", "TOTAL GÉNÉRAL", total_general, "", ""])
        
        # Style de la ligne de total
        total_row = ws.max_row
        ws[f'B{total_row}'].font = Font(bold=True)
        ws[f'C{total_row}'].font = Font(bold=True)
        ws[f'C{total_row}'].number_format = '#,##0'

        # Format des montants
        for row in range(5, total_row):
            ws[f'C{row}'].number_format = '#,##0'

        # Ajuster largeur colonnes
        for i, column_cells in enumerate(ws.columns, 1):
            max_length = 0
            for cell in column_cells:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            # Limiter à 50 caractères max
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[get_column_letter(i)].width = adjusted_width

        # Response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=rapport_financier.xlsx'

        wb.save(response)
        return response

    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return JsonResponse({
            "success": False,
            "message": str(e)
        })

def generate_dette_excel(request):

    id_annee = request.GET.get('annee')
    id_classe = request.GET.get('classe')
    id_trimestre = request.GET.get('trimestre')

    if not id_annee or not id_classe:
        return JsonResponse({
            "success": False,
            "message": "Année et classe obligatoires"
        })

    # 🔥 TRI ALPHABETIQUE ICI
    inscriptions = Eleve_inscription.objects.select_related(
        'id_eleve',
        'id_classe',
        'id_classe__classe_id'
    ).filter(
        id_annee=id_annee,
        id_classe=id_classe,
        status=True
    ).order_by('id_eleve__nom', 'id_eleve__prenom')

    if id_trimestre:
        inscriptions = inscriptions.filter(id_trimestre=id_trimestre)

    wb = Workbook()
    ws = wb.active
    ws.title = "Rapport Dettes"

    # Titre
    ws.merge_cells('A1:G1')
    ws['A1'] = "RAPPORT DES DETTES"
    ws['A1'].font = Font(size=14, bold=True)
    ws['A1'].alignment = Alignment(horizontal="center")

    headers = [
        "Élève",
        "Variable",
        "À payer",
        "Payé",
        "Reste",
        "Pénalité",
        "Total"
    ]

    ws.append(headers)

    for cell in ws[2]:
        cell.font = Font(bold=True)

    row_index = 3

    for ins in inscriptions:

        eleve = ins.id_eleve

        paiements = Paiement.objects.filter(
            id_eleve=eleve,
            id_classe_active=id_classe,
            id_annee=id_annee
        )

        variables = VariablePrix.objects.filter(
            id_classe_active=id_classe,
            id_annee=id_annee,
            id_variable__estObligatoire=True
        ).select_related('id_variable')

        total_dette_eleve = 0

        for vp in variables:

            variable = vp.id_variable
            montant_a_payer = vp.prix

            montant_paye = paiements.filter(
                id_variable=variable
            ).aggregate(total=Sum('montant'))['total'] or 0

            reste = max(montant_a_payer - montant_paye, 0)

            penalite = 0
            if reste > 0:
                penalite = PenaliteConfig.objects.filter(
                    id_variable=variable,
                    id_classe_active=id_classe,
                    actif=True
                ).aggregate(total=Sum('valeur'))['total'] or 0

            total = reste + penalite

            if total > 0:

                ws.append([
                    f"{eleve.nom} {eleve.prenom}",
                    variable.variable,
                    montant_a_payer,
                    montant_paye,
                    reste,
                    penalite,
                    total
                ])

                total_dette_eleve += total

        # Ligne total élève
        if total_dette_eleve > 0:
            ws.append([
                "",
                "TOTAL ELEVE",
                "",
                "",
                "",
                "",
                total_dette_eleve
            ])

            ws[f'B{ws.max_row}'].font = Font(bold=True)
            ws[f'G{ws.max_row}'].font = Font(bold=True)

    # Ajuster largeur colonnes (safe merged)
    for i, column_cells in enumerate(ws.columns, 1):
        max_length = 0
        for cell in column_cells:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass

        ws.column_dimensions[get_column_letter(i)].width = max_length + 2

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

    response['Content-Disposition'] = 'attachment; filename=rapport_dettes.xlsx'

    wb.save(response)

    return response

def generate_situation_excel(request):

    date_debut = request.GET.get('date_debut')
    date_fin = request.GET.get('date_fin')

    paiements = Paiement.objects.select_related(
        'id_eleve',
        'id_variable',
        'id_compte'
    ).filter(
        status=True,
        is_rejected=False
    )

    # ======================
    # FILTRE DATE
    # ======================
    if date_debut and date_fin:
        paiements = paiements.filter(
            date_saisie__range=[date_debut, date_fin]
        )
    else:
        today = date.today()
        paiements = paiements.filter(date_saisie=today)
        date_debut = today
        date_fin = today

    # ======================
    # TRI ALPHABETIQUE
    # ======================
    paiements = paiements.order_by(
        'id_eleve__nom',
        'id_eleve__prenom'
    )

    total_general = paiements.aggregate(
        total=Sum('montant')
    )['total'] or 0

    # ======================
    # EXCEL
    # ======================
    wb = Workbook()
    ws = wb.active
    ws.title = "Situation Journalière"

    # TITRE
    ws.merge_cells('A1:E1')
    ws['A1'] = f"SITUATION DES PAIEMENTS DU {date_debut} AU {date_fin}"
    ws['A1'].font = Font(size=14, bold=True)
    ws['A1'].alignment = Alignment(horizontal="center")

    headers = [
        "Élève",
        "Variable",
        "Montant",
        "Date paiement",
        "Compte utilisé"
    ]

    ws.append(headers)

    for cell in ws[2]:
        cell.font = Font(bold=True)

    # ======================
    # DONNEES
    # ======================
    for p in paiements:

        ws.append([
            str(p.id_eleve),
            str(p.id_variable),
            p.montant,
            p.date_paie.strftime("%Y-%m-%d") if p.date_paie else "",
            str(p.id_compte) if p.id_compte else "-"
        ])

    # ======================
    # TOTAL
    # ======================
    ws.append([
        "",
        "TOTAL",
        total_general,
        "",
        ""
    ])

    ws[f'B{ws.max_row}'].font = Font(bold=True)
    ws[f'C{ws.max_row}'].font = Font(bold=True)

    # ======================
    # AUTO WIDTH SAFE
    # ======================
    for i, column_cells in enumerate(ws.columns, 1):
        max_length = 0
        for cell in column_cells:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass

        ws.column_dimensions[get_column_letter(i)].width = max_length + 2

    # ======================
    # RESPONSE
    # ======================
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

    filename = f"situation_paiements_{date_debut}_au_{date_fin}.xlsx"

    response['Content-Disposition'] = f'attachment; filename={filename}'

    wb.save(response)

    return response

def export_dashboard_excel(request):
    type_stat = request.GET.get('type')
    annee_id = request.GET.get('annee')
    classe_id = request.GET.get('classe')
    eleve_id = request.GET.get('eleve')
    variable_id = request.GET.get('variable')
    trimestre_id = request.GET.get('trimestre')

    if not annee_id:
        return JsonResponse({"success": False, "message": "Année obligatoire"})

    # ======================
    # Classe active
    # ======================
    classe_active = None
    if classe_id and classe_id.isdigit():
        classe_active = Classe_active.objects.filter(id_classe_active=int(classe_id)).first()

    # ======================
    # Élèves
    # ======================
    eleves = Eleve_inscription.objects.filter(id_annee_id=annee_id, status=True)
    if classe_active:
        eleves = eleves.filter(
            id_classe_id=classe_active.id_classe_active,
            id_campus_id=classe_active.id_campus,
            id_classe_cycle_id=classe_active.cycle_id
        )
    if eleve_id and eleve_id.isdigit():
        eleves = eleves.filter(id_eleve_id=int(eleve_id))

    # ======================
    # Variables
    # ======================
    variables_prix = VariablePrix.objects.filter(id_annee_id=annee_id)
    if classe_active:
        variables_prix = variables_prix.filter(id_classe_active=classe_active.id_classe_active)
    if variable_id and variable_id.isdigit():
        variables_prix = variables_prix.filter(id_variable_id=int(variable_id))
    if trimestre_id and trimestre_id.isdigit():
        variables_prix = variables_prix.filter(id_annee_trimestre_id=int(trimestre_id))

    # ======================
    # Fonction utilitaire
    # ======================
    def get_classe_full_from_vp(vp):
        c = vp.id_classe_active
        return f"{c.id_campus.campus} - {c.cycle_id.cycle_id.cycle} - {c.classe_id.classe} {c.groupe or ''}"

    # ======================
    # Construction des données
    # ======================
    data = []
    title = ""

    if type_stat == "transactions":
        title = "Total Paiements par Variable"
        for vp in variables_prix:
            qs = Paiement.objects.filter(
                id_variable_id=vp.id_variable_id,
                id_annee_id=annee_id,
                status=True,
                is_rejected=False
            )
            if classe_active:
                qs = qs.filter(id_classe_active_id=classe_active.id_classe_active)
            data.append([get_classe_full_from_vp(vp), '-', vp.id_variable.variable, qs.count()])

    elif type_stat == "paye":
        title = "Montant Payé par Élève"
        for vp in variables_prix:
            for e in eleves:
                paiements_qs = Paiement.objects.filter(
                    id_variable_id=vp.id_variable_id,
                    id_eleve_id=e.id_eleve_id,
                    id_annee_id=annee_id,
                    status=True,
                    is_rejected=False
                )
                if classe_active:
                    paiements_qs = paiements_qs.filter(id_classe_active_id=classe_active.id_classe_active)
                total_montant = paiements_qs.aggregate(total=Sum('montant'))['total'] or 0
                if total_montant > 0:
                    data.append([get_classe_full_from_vp(vp), f"{e.id_eleve.nom} {e.id_eleve.prenom}", vp.id_variable.variable, total_montant])

    elif type_stat == "rejet":
        title = "Paiements Rejetés"
        for vp in variables_prix:
            qs = Paiement.objects.filter(
                id_variable_id=vp.id_variable_id,
                id_annee_id=annee_id,
                is_rejected=True
            )
            if classe_active:
                qs = qs.filter(id_classe_active_id=classe_active.id_classe_active)
            data.append([get_classe_full_from_vp(vp), '-', vp.id_variable.variable, qs.count()])

    elif type_stat == "attendu":
        title = "Montant Attendu par Variable"
        for vp in variables_prix:
            total_attendu = 0
            prix = vp.prix
            for e in eleves:
                reduction = Eleve_reduction_prix.objects.filter(
                    id_variable_id=vp.id_variable_id,
                    id_eleve_id=e.id_eleve_id,
                    id_annee_id=annee_id
                ).first()
                if reduction:
                    total_attendu += prix - (prix * reduction.pourcentage / 100)
                else:
                    total_attendu += prix
            data.append([get_classe_full_from_vp(vp), '-', vp.id_variable.variable, total_attendu])

    elif type_stat == "reste":
        title = "Reste à Payer par Variable"
        for vp in variables_prix:
            reste_global = 0
            prix = vp.prix
            for e in eleves:
                reduction = Eleve_reduction_prix.objects.filter(
                    id_variable_id=vp.id_variable_id,
                    id_eleve_id=e.id_eleve_id,
                    id_annee_id=annee_id
                ).first()
                attendu = prix - (prix * reduction.pourcentage / 100) if reduction else prix
                total_paye = Paiement.objects.filter(
                    id_variable_id=vp.id_variable_id,
                    id_eleve_id=e.id_eleve_id,
                    id_annee_id=annee_id,
                    status=True,
                    is_rejected=False
                ).aggregate(total=Sum('montant'))['total'] or 0
                reste_global += max(attendu - total_paye, 0)
            data.append([get_classe_full_from_vp(vp), '-', vp.id_variable.variable, reste_global])

    elif type_stat == "dette":
        title = "Élèves en Dette"
        for vp in variables_prix:
            prix = vp.prix
            for e in eleves:
                reduction = Eleve_reduction_prix.objects.filter(
                    id_variable_id=vp.id_variable_id,
                    id_eleve_id=e.id_eleve_id,
                    id_annee_id=annee_id
                ).first()
                attendu = prix - (prix * reduction.pourcentage / 100) if reduction else prix
                total_paye = Paiement.objects.filter(
                    id_variable_id=vp.id_variable_id,
                    id_eleve_id=e.id_eleve_id,
                    id_annee_id=annee_id,
                    status=True,
                    is_rejected=False
                ).aggregate(total=Sum('montant'))['total'] or 0
                reste = attendu - total_paye
                if reste > 0:
                    data.append([get_classe_full_from_vp(vp), f"{e.id_eleve.nom} {e.id_eleve.prenom}", vp.id_variable.variable, reste])

    # ======================
    # Création Excel
    # ======================

    wb = Workbook()
    ws = wb.active
    ws.title = "Dashboard"

    ws.merge_cells('A1:D1')
    ws['A1'] = title
    ws['A1'].font = Font(size=14, bold=True)
    ws['A1'].alignment = Alignment(horizontal="center")

    headers = ["Classe", "Élève", "Variable", "Total"]
    ws.append(headers)
    for cell in ws[2]:
        cell.font = Font(bold=True)

    total_general = 0
    for row in data:
        total_general += row[3] or 0
        ws.append(row)

    ws.append(["", "", "TOTAL GENERAL", total_general])
    ws[f'C{ws.max_row}'].font = Font(bold=True)
    ws[f'D{ws.max_row}'].font = Font(bold=True)

    for i, column_cells in enumerate(ws.columns, 1):
        max_length = max((len(str(cell.value)) for cell in column_cells if cell.value), default=0)
        ws.column_dimensions[get_column_letter(i)].width = max_length + 2

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=dashboard_{type_stat}.xlsx'
    wb.save(response)
    return response
