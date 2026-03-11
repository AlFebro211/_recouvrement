from app.forms.recouvrement_forms import PaiementUpdateForm

# Imports depuis create_base (utilitaires Django + forms + logger)
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.views.decorators.csrf import csrf_protect, csrf_exempt
import logging
logger = logging.getLogger(__name__)
from app.forms.recouvrement_forms import (
    VariableDateButoireForm, PaiementForm, VariableCategorieForm,
    VariableForm, BanqueForm, CompteForm, VariablePrixForm,
    VariableDerogationForm, VariableReductionForm, PenaliteForm,
    CategorieOperationForm, OperationCaisseForm,
)

# Imports des modèles
from app.models import (
    Banque, VariableCategorie, Annee, Classe_active, Paiement,
    Compte, VariablePrix, Eleve_reduction_prix, Eleve_inscription,
    Annee_trimestre, VariableDatebutoire, VariableDerogation,
    PenaliteConfig, Variable, Campus, Classe, Classe_cycle_actif,
    CategorieOperation, OperationCaisse, Eleve,
)

# Imports Django
from django.db.models import Q, Sum
from django.db.models.functions import Coalesce
from django.http import JsonResponse, HttpResponse

# Imports Python standard
from datetime import date, datetime
import io
from io import BytesIO

# Imports pour le fichier Excel
from app.views.invoice_paiement import build_pdf_header
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Imports pour PDF avec ReportLab
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT



def get_existing_derogation_reduction(request):
    """
    Returns existing derogation dates, reduction percentages, and date_butoire values
    for a student in a given class/year. Used to pre-fill form fields.
    """
    id_annee = request.GET.get('id_annee')
    id_campus = request.GET.get('id_campus')
    id_cycle = request.GET.get('id_cycle')
    id_classe_active = request.GET.get('id_classe_active')
    id_eleve = request.GET.get('id_eleve')

    if not all([id_annee, id_classe_active]):
        return JsonResponse({'success': False, 'error': 'Paramètres manquants'}, status=400)

    try:
        result = {}

        # Date butoire par variable (pour toute la classe)
        dates_butoire = VariableDatebutoire.objects.filter(
            id_annee_id=id_annee,
            id_classe_active_id=id_classe_active,
        )
        for db in dates_butoire:
            var_id = str(db.id_variable_id)
            if var_id not in result:
                result[var_id] = {}
            result[var_id]['date_butoire'] = db.date_butoire.strftime('%Y-%m-%d') if db.date_butoire else None

        # Derogation par élève + variable
        if id_eleve:
            derogations = VariableDerogation.objects.filter(
                id_annee_id=id_annee,
                id_classe_active_id=id_classe_active,
                id_eleve_id=id_eleve,
            )
            for d in derogations:
                var_id = str(d.id_variable_id)
                if var_id not in result:
                    result[var_id] = {}
                result[var_id]['date_derogation'] = d.date_derogation.strftime('%Y-%m-%d') if d.date_derogation else None

            # Reduction par élève + variable
            reductions = Eleve_reduction_prix.objects.filter(
                id_annee_id=id_annee,
                id_classe_active_id=id_classe_active,
                id_eleve_id=id_eleve,
            )
            for r in reductions:
                var_id = str(r.id_variable_id)
                if var_id not in result:
                    result[var_id] = {}
                result[var_id]['pourcentage'] = float(r.pourcentage)

        return JsonResponse({'success': True, 'data': result})

    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


def get_banques(request):
    banques = Banque.objects.all().values('id_banque', 'banque')
    return JsonResponse({'banques': list(banques)})


def get_categories(request):
    categories = VariableCategorie.objects.all().values('id_variable_categorie', 'nom')
    return JsonResponse({'categories': list(categories)})


@csrf_protect
def store_annee_session(request):
    if request.method == 'POST':
        try:
            annee_id = request.POST.get('id_annee')
            if not annee_id:
                return JsonResponse({'success': False, 'error': 'Aucune année sélectionnée.'})
            request.session['selected_annee_id'] = int(annee_id)
            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    return JsonResponse({'success': False, 'error': 'Méthode non autorisée.'})



def get_classes_actives(request, annee_id):
    try:
        
        if not Annee.objects.filter(id_annee=annee_id).exists():
            return JsonResponse({
                'success': False,
                'error': 'Année scolaire non trouvée.'
            }, status=404)

        classes = (
            Classe_active.objects.filter(
                id_annee=annee_id,
                is_active=True,
                eleve_inscription__status=1  
            )
            .select_related('id_campus', 'cycle_id__cycle_id', 'classe_id')
            .values(
                'id_classe_active',
                'id_campus__campus',
                'cycle_id__cycle_id__cycle',
                'classe_id__classe',
                'groupe',
                'id_campus__id_campus',
                'cycle_id__id_cycle_actif',
            )
            .distinct() 
        )

        formatted_classes = [
            {
                'id_classe_active': classe['id_classe_active'],
                'campus_nom': classe['id_campus__campus'],
                'cycle_nom': classe['cycle_id__cycle_id__cycle'],
                'classe_nom': classe['classe_id__classe'],
                'groupe': classe['groupe'] or '',
                'id_campus': classe['id_campus__id_campus'],
                'id_cycle': classe['cycle_id__id_cycle_actif'],
            }
            for classe in classes
        ]

        if not formatted_classes:
            return JsonResponse({
                'success': True,
                'classes': [],
                'message': 'Aucune classe active avec des inscrits trouvée pour cette année scolaire.'
            })

        return JsonResponse({
            'success': True,
            'classes': formatted_classes,
        })

    except Exception as e:
        logger.error(f"Erreur dans get_classes_actives pour annee_id={annee_id}: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': 'Une erreur est survenue lors de la récupération des classes.'
        }, status=500)



def get_classes_actives_avec_paiement(request, annee_id):
    try:
        if not Annee.objects.filter(id_annee=annee_id).exists():
            return JsonResponse({
                'success': False,
                'error': 'Année scolaire non trouvée.'
            }, status=404)

        classes = (
            Classe_active.objects.filter(
                id_annee=annee_id,
                is_active=True,
                paiement__id_annee=annee_id,
                paiement__status = 1
            )
            .select_related('id_campus', 'cycle_id__cycle_id', 'classe_id')
            .values(
                'id_classe_active',
                'id_campus__campus',
                'cycle_id__cycle_id__cycle',
                'classe_id__classe',
                'groupe',
                'id_campus__id_campus',
                'cycle_id__id_cycle_actif',
            )
            .distinct()
        )

        formatted_classes = [
            {
                'id_classe_active': classe['id_classe_active'],
                'campus_nom': classe['id_campus__campus'],
                'cycle_nom': classe['cycle_id__cycle_id__cycle'],
                'classe_nom': classe['classe_id__classe'],
                'groupe': classe['groupe'] or '',
                'id_campus': classe['id_campus__id_campus'],
                'id_cycle': classe['cycle_id__id_cycle_actif'],
            }
            for classe in classes
        ]

        if not formatted_classes:
            return JsonResponse({
                'success': True,
                'classes': [],
                'message': 'Aucune classe active avec des paiements trouvée pour cette année scolaire.'
            })

        return JsonResponse({
            'success': True,
            'classes': formatted_classes,
        })

    except Exception as e:
        logger.error(f"Erreur dans get_classes_actives pour annee_id={annee_id}: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': 'Une erreur est survenue lors de la récupération des classes.'
        }, status=500)



def get_comptes_banque(request, id_banque):
    try:
        comptes = Compte.objects.filter(id_banque_id=id_banque).values('id_compte', 'compte')
        formatted_comptes = [
            {
                'id_compte': compte['id_compte'],
                'compte': compte['compte']
            } for compte in comptes
        ]
        return JsonResponse({'success': True, 'comptes': formatted_comptes})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)



#
def get_all_paiement_soumises(request):
    
    form = VariableDateButoireForm()
    all_paiement_soumises = Paiement.objects.all()
    return render(request, 'recouvrement/index_recouvrement.html', {
        'paiement_list':all_paiement_soumises,
        'form_paiement_validation': form,
        'form_type': 'validation_form',
       

    })



@csrf_exempt
def get_paiements_submitted(request):

    if request.method == "GET":
        try:
            id_campus = request.GET.get("id_campus")
            id_cycle_actif = request.GET.get("id_cycle")
            id_classe_active = request.GET.get("id_classe_active")
            id_annee = request.GET.get("id_annee")

            if not all([id_campus, id_cycle_actif, id_classe_active, id_annee]):
                return JsonResponse({
                    "success": False,
                    "error": "Paramètres requis manquants."
                }, status=400)

            paiements = Paiement.objects.filter(
                id_campus_id=id_campus,
                id_cycle_actif_id=id_cycle_actif,
                id_classe_active_id=id_classe_active,
                id_annee_id=id_annee,
                status = 0,
                is_rejected = 0
                
            ).select_related("id_eleve", "id_variable")

            data = []
            for p in paiements:
                data.append({
                    "id_paiement": p.id_paiement,
                    "id_variable": p.id_variable.id_variable,
                    "variable": p.id_variable.variable,
                    "montant": p.montant,
                    "bordereau": p.bordereau.name if p.bordereau else "",
                    "status": p.status,
                    "eleve_nom": p.id_eleve.nom,
                    "eleve_prenom": p.id_eleve.prenom,
                })

            return JsonResponse({"success": True, "data": data}, safe=False)

        except Exception as e:
            return JsonResponse({
                "success": False,
                "error": f"Erreur serveur : {str(e)}"
            }, status=500)

    return JsonResponse({"success": False, "error": "Méthode non autorisée"}, status=405)



@csrf_exempt
def get_paiements_validated(request):
  
    if request.method == "GET":
        try:
            id_campus = request.GET.get("id_campus")
            id_cycle_actif = request.GET.get("id_cycle")
            id_classe_active = request.GET.get("id_classe_active")
            id_annee = request.GET.get("id_annee")

            if not all([id_campus, id_cycle_actif, id_classe_active, id_annee]):
                return JsonResponse({
                    "success": False,
                    "error": "Paramètres requis manquants."
                }, status=400)

            paiements = Paiement.objects.filter(
                id_campus_id=id_campus,
                id_cycle_actif_id=id_cycle_actif,
                id_classe_active_id=id_classe_active,
                id_annee_id=id_annee,
                status = 1,
                
            ).select_related("id_eleve", "id_variable")

            data = []
            for p in paiements:
                data.append({
                    "id_paiement": p.id_paiement,
                    "id_eleve": p.id_eleve.id_eleve,
                    "id_variable": p.id_variable.id_variable,
                    "variable": p.id_variable.variable,
                    "montant": p.montant,
                    "bordereau": p.bordereau.name if p.bordereau else "",
                    "status": p.status,
                    "eleve_nom": p.id_eleve.nom,
                    "eleve_prenom": p.id_eleve.prenom,
                })

            return JsonResponse({"success": True, "data": data}, safe=False)

        except Exception as e:
            return JsonResponse({
                "success": False,
                "error": f"Erreur serveur : {str(e)}"
            }, status=500)

    return JsonResponse({"success": False, "error": "Méthode non autorisée"}, status=405)

def get_paiements_for_add_page(request):
    """
    Retourne les paiements filtrés pour la page d'ajout,
    seulement ceux avec status=True
    """
    id_annee = request.GET.get('id_annee')
    id_classe = request.GET.get('id_classe_active')
    id_eleve = request.GET.get('id_eleve')

    paiements = Paiement.objects.filter(status=True)

    if id_annee:
        paiements = paiements.filter(id_annee_id=id_annee)
    if id_classe:
        paiements = paiements.filter(id_classe_active_id=id_classe)
    if id_eleve:
        paiements = paiements.filter(id_eleve_id=id_eleve)

    data = []
    for p in paiements:
        data.append({
            "id": p.id_paiement,
            "eleve": str(p.id_eleve),
            "variable": str(p.id_variable),
            "montant": p.montant,
            "bordereau": p.bordereau.url if p.bordereau else None,
            "date_paie": p.date_paie.strftime("%Y-%m-%d") if p.date_paie else None,
            "status": p.status,
            "is_rejected": p.is_rejected
        })

    return JsonResponse({"success": True, "data": data})


def get_pupils_with_unpaid(request):
    """
    Returns only students who have unpaid variables for the given trimester.
    Takes trimestre (id_annee_trimestre) as a parameter.
    Excludes students who have fully paid all variables.
    """
    id_annee = request.GET.get('id_annee')
    id_campus = request.GET.get('id_campus')
    id_cycle = request.GET.get('id_cycle')
    id_classe = request.GET.get('id_classe_active')
    id_trimestre = request.GET.get('id_trimestre')

    if not all([id_annee, id_campus, id_cycle, id_classe, id_trimestre]):
        return JsonResponse({'success': False, 'error': 'Paramètres manquants'}, status=400)

    try:
        # Variables prix pour ce trimestre + classe
        variables_prix = VariablePrix.objects.filter(
            id_annee_trimestre_id=id_trimestre,
            id_annee_id=id_annee,
            id_campus_id=id_campus,
            id_cycle_actif_id=id_cycle,
            id_classe_active_id=id_classe
        ).select_related('id_variable')

        if not variables_prix.exists():
            return JsonResponse({'success': True, 'data': [], 'count': 0,
                                 'message': 'Aucune variable définie pour ce trimestre.'})

        # Élèves inscrits dans cette classe
        inscriptions = Eleve_inscription.objects.filter(
            id_annee_id=id_annee,
            id_campus_id=id_campus,
            id_classe_cycle_id=id_cycle,
            id_classe_id=id_classe,
            status=1
        ).select_related('id_eleve').distinct()

        pupils_data = []

        for ins in inscriptions:
            eleve = ins.id_eleve
            has_unpaid = False
            total_reste = 0

            for vp in variables_prix:
                montant = vp.prix

                # Réduction éventuelle
                reduction = Eleve_reduction_prix.objects.filter(
                    id_eleve=eleve,
                    id_variable=vp.id_variable,
                    id_annee_id=id_annee,
                    id_campus_id=id_campus,
                    id_cycle_actif_id=id_cycle,
                    id_classe_active_id=id_classe
                ).first()

                if reduction:
                    montant -= montant * reduction.pourcentage / 100

                # Somme déjà payée
                paye = Paiement.objects.filter(
                    id_eleve=eleve,
                    id_variable=vp.id_variable,
                    id_annee_id=id_annee,
                    status=True
                ).aggregate(total=Sum('montant'))['total'] or 0

                reste = max(0, montant - paye)
                if reste > 0:
                    has_unpaid = True
                    total_reste += reste

            # Only include students with unpaid balances
            if has_unpaid:
                pupils_data.append({
                    'id_eleve': eleve.id_eleve,
                    'nom_complet': f"{eleve.nom} {eleve.prenom}",
                    'reste_a_payer': total_reste,
                })

        return JsonResponse({
            'success': True,
            'data': pupils_data,
            'count': len(pupils_data)
        })

    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@csrf_exempt
def get_pupils_registred_classe(request):
    id_annee = request.GET.get('id_annee')
    id_campus = request.GET.get('id_campus')
    id_cycle = request.GET.get('id_cycle')
    id_classe = request.GET.get('id_classe_active')

    if not all([id_annee, id_campus, id_cycle, id_classe]):
        return JsonResponse({'error': 'Paramètres manquants'}, status=400)

    try:
        from django.utils.timezone import now
        today = now().date()

        # 🔹 Trimestre en cours
        trimestre_en_cours = Annee_trimestre.objects.filter(
            id_annee=id_annee,
            id_campus=id_campus,
            id_cycle=id_cycle,
            id_classe=id_classe,
            # debut__lte=today,
            # fin__gte=today,
            etat_trimestre='En attente'
        ).first()

        if not trimestre_en_cours:
            return JsonResponse({'success': True, 'data': [], 'count': 0})

        # 🔹 Variables définies pour la classe (source mobile)
        variables_prix = VariablePrix.objects.filter(
            id_annee=id_annee,
            id_campus=id_campus,
            id_cycle_actif=id_cycle,
            id_classe_active=id_classe
        ).select_related('id_variable')

        # 🔹 Élèves inscrits
        inscriptions = Eleve_inscription.objects.filter(
            id_annee=id_annee,
            id_campus=id_campus,
            id_classe_cycle=id_cycle,
            id_classe=id_classe,
            status=1
        ).select_related('id_eleve').distinct()

        pupils_data = []

        for ins in inscriptions:
            eleve = ins.id_eleve
            total_a_payer = 0
            total_paye = 0

            for vp in variables_prix:
                montant = vp.prix

                # 🔻 Réduction éventuelle
                reduction = Eleve_reduction_prix.objects.filter(
                    id_eleve=eleve,
                    id_variable=vp.id_variable,
                    id_annee=id_annee,
                    id_campus=id_campus,
                    id_cycle_actif=id_cycle,
                    id_classe_active=id_classe
                ).first()

                if reduction:
                    montant -= montant * reduction.pourcentage / 100

                total_a_payer += montant

                paye = Paiement.objects.filter(
                    id_eleve=eleve,
                    id_variable=vp.id_variable
                ).aggregate(total=Sum('montant'))['total'] or 0

                total_paye += paye

            # 🔥 IMPORTANT : afficher seulement les débiteurs
            if total_paye < total_a_payer:
                pupils_data.append({
                    'id_eleve': eleve.id_eleve,
                    'nom_complet': f"{eleve.nom} {eleve.prenom}",
                    'reste_a_payer': total_a_payer - total_paye,
                    'trimestre': trimestre_en_cours.trimestre.trimestre
                })

        return JsonResponse({
            'success': True,
            'data': pupils_data,
            'count': len(pupils_data)
        })

    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)



# @csrf_exempt
# def get_pupils_registred_classe(request):
#     id_annee = request.GET.get('id_annee')
#     id_campus = request.GET.get('id_campus')
#     id_cycle = request.GET.get('id_cycle')
#     id_classe = request.GET.get('id_classe_active')

#     if not all([id_annee, id_campus, id_cycle, id_classe]):
#         return JsonResponse({'error': 'Paramètres manquants'}, status=400)
    

#     try:
#         inscriptions = Eleve_inscription.objects.filter(
#             Q(id_annee=id_annee) &
#             Q(id_campus=id_campus) &
#             Q(id_classe_cycle=id_cycle) &
#             Q(id_classe=id_classe), 
#             Q(status=1)
#         ).select_related('id_eleve').order_by('id_eleve')

#         seen_eleve_ids = set()
#         pupils_data = []
#         for inscription in inscriptions:
#             eleve = inscription.id_eleve
#             if eleve.id_eleve not in seen_eleve_ids:
#                 seen_eleve_ids.add(eleve.id_eleve)
#                 pupils_data.append({
#                     'id_eleve': eleve.id_eleve,
#                     'status': inscription.status, 
#                     'redoublement': inscription.redoublement,
#                     'nom_complet': f"{eleve.nom} {eleve.prenom}"
#                 })

#         return JsonResponse({
#             'success': True,
#             'data': pupils_data,
#             'count': len(pupils_data)
#         })

#     except Exception as e:
#         return JsonResponse({
#             'error': str(e),
#             'success': False
#         }, status=500)


# def get_variables_restant_a_payer(request):
#     eleve_id = request.GET.get('id_eleve')
#     annee_id = request.GET.get('id_annee')
#     campus_id = request.GET.get('id_campus')
#     cycle_id = request.GET.get('id_cycle')
#     classe_id = request.GET.get('id_classe')

#     variables_prix = VariablePrix.objects.filter(
#         id_annee_id=annee_id,
#         id_campus_id=campus_id,
#         id_cycle_actif_id=cycle_id,
#         id_classe_active_id=classe_id
#     ).select_related('id_variable')

#     result = []

#     for vp in variables_prix:
#         variable = vp.id_variable
#         montant_max = vp.prix

#         reduction = Eleve_reduction_prix.objects.filter(
#             id_eleve_id=eleve_id,
#             id_variable_id=variable.id_variable,
#             id_annee_id=annee_id,
#             id_campus_id=campus_id,
#             id_cycle_actif_id=cycle_id,
#             id_classe_active_id=classe_id
#         ).first()

#         if reduction:
#             montant_max -= (montant_max * reduction.pourcentage / 100)

#         total_paye = Paiement.objects.filter(
#             id_eleve_id=eleve_id,
#             id_variable_id=variable.id_variable,
#             id_annee_id=annee_id,
#             id_campus_id=campus_id,
#             id_cycle_actif_id=cycle_id,
#             id_classe_active_id=classe_id
#         ).aggregate(total=Sum('montant'))['total'] or 0

#         reste = max(0, montant_max - total_paye)

#         if reste > 0:
#             result.append({
#                 'id_variable': variable.id_variable,
#                 'nom_variable': variable.variable,
#                 'montant_total': montant_max,
#                 'total_deja_paye': total_paye,
#                 'reste_a_payer': reste,
#                 'reduction': reduction.pourcentage if reduction else 0
#             })

#     return JsonResponse({'success': True, 'variables': result})


def get_variables_restant_a_payer(request):
    eleve_id = request.GET.get('id_eleve')
    annee_id = request.GET.get('id_annee')
    campus_id = request.GET.get('id_campus')
    cycle_id = request.GET.get('id_cycle')
    classe_id = request.GET.get('id_classe')

    # 🔹 Pour l'instant on récupère même si le trimestre est "En attente"
    annee_trimestre = Annee_trimestre.objects.filter(
        id_annee_id=annee_id,
        id_campus_id=campus_id,
        id_cycle_id=cycle_id,
        id_classe_id=classe_id,
        # etat_trimestre="En cours"
        etat_trimestre="En attente"
    ).select_related('trimestre').first()

    if not annee_trimestre:
        return JsonResponse({
            'success': False,
            'error': 'Aucun trimestre trouvé pour cette classe'
        }, status=400)

    # 🔹 Récupérer toutes les variables liées à ce trimestre
    variables_prix = VariablePrix.objects.filter(
        id_annee_trimestre=annee_trimestre
    ).select_related('id_variable', 'id_variable__id_variable_categorie')


    result = []

    for vp in variables_prix:
        variable = vp.id_variable
        montant_max = vp.prix

        # 🔹 Appliquer réduction si elle existe
        reduction = Eleve_reduction_prix.objects.filter(
            id_eleve_id=eleve_id,
            id_variable_id=variable.id_variable,
            id_annee_id=annee_id,
            id_campus_id=campus_id,
            id_cycle_actif_id=cycle_id,
            id_classe_active_id=classe_id
        ).first()

        if reduction:
            montant_max -= montant_max * reduction.pourcentage / 100

        # 🔹 Somme déjà payée
        total_paye = Paiement.objects.filter(
            id_eleve_id=eleve_id,
            id_variable_id=variable.id_variable,
            id_annee_id=annee_id,
            id_campus_id=campus_id,
            id_cycle_actif_id=cycle_id,
            id_classe_active_id=classe_id
        ).aggregate(total=Sum('montant'))['total'] or 0

        # 🔹 Calculer le reste à payer
        reste = max(0, montant_max - total_paye)

        if reste > 0:
            result.append({
                'id_variable': variable.id_variable,
                'nom_variable': variable.variable,
                'categorie': variable.id_variable_categorie.nom,
                'montant_total': montant_max,
                'total_deja_paye': total_paye,
                'reste_a_payer': reste,
                'trimestre': annee_trimestre.trimestre.trimestre,
                'id_trimestre': annee_trimestre.trimestre.id_trimestre
            })

    return JsonResponse({
        'success': True,
        'annee_trimestre': {
            'id': annee_trimestre.trimestre.id_trimestre,
            'trimestre': annee_trimestre.trimestre.trimestre,
            'etat': annee_trimestre.etat_trimestre,
            'debut': annee_trimestre.debut,
            'fin': annee_trimestre.fin
        },
        'variables': result
    })



def get_trimestres_by_classe_active(request):
    classe_active_id = request.GET.get('id_classe_active')
    annee_id = request.GET.get('id_annee')
    campus_id = request.GET.get('id_campus')
    cycle_id = request.GET.get('id_cycle')

    if not all([classe_active_id, annee_id, campus_id, cycle_id]):
        return JsonResponse({'success': False, 'error': 'Paramètres manquants'})

    trimestres_qs = Annee_trimestre.objects.filter(
        id_annee_id=annee_id,
        id_classe_id=classe_active_id,
        id_campus_id=campus_id,
        id_cycle_id=cycle_id
    ).select_related('trimestre').order_by('trimestre__trimestre')

    data = [
        {
            'id': t.id_trimestre,  # ✅ IMPORTANT (plus trimestre.id)
            'nom': t.trimestre.trimestre,
            'etat': t.etat_trimestre
        }
        for t in trimestres_qs
    ]

    return JsonResponse({'success': True, 'trimestres': data})



def get_variables_by_trimestre(request):
    # 🔹 Paramètres
    annee_trimestre_id = request.GET.get('id_trimestre')  # id_annee_trimestre
    classe_id = request.GET.get('id_classe')
    annee_id = request.GET.get('id_annee')
    campus_id = request.GET.get('id_campus')
    cycle_id = request.GET.get('id_cycle')

    if not all([annee_trimestre_id, classe_id, annee_id, campus_id, cycle_id]):
        return JsonResponse({
            'success': False,
            'error': 'Paramètres manquants'
        }, status=400)

    # 🔹 Récupérer le trimestre
    annee_trimestre = Annee_trimestre.objects.filter(
        id_trimestre=annee_trimestre_id
    ).select_related('trimestre').first()

    if not annee_trimestre:
        return JsonResponse({
            'success': False,
            'error': 'Trimestre invalide'
        }, status=400)

    # 🔹 Récupérer les VariablePrix liées
    variable_prix_qs = VariablePrix.objects.filter(
        id_annee_trimestre_id=annee_trimestre_id,
        id_classe_active_id=classe_id,
        id_annee_id=annee_id,
        id_campus_id=campus_id,
        id_cycle_actif_id=cycle_id
    ).select_related('id_variable', 'id_variable__id_variable_categorie') \
     .order_by('id_variable_id')

    # 🔹 Préparer la liste des variables avec prix
    variables_data = [
        {
            'id_variable': vp.id_variable.id_variable,
            'nom_variable': vp.id_variable.variable,
            'categorie': vp.id_variable.id_variable_categorie.nom,
            'prix': vp.prix
        }
        for vp in variable_prix_qs
    ]

    return JsonResponse({
        'success': True,
        'annee_trimestre': {
            'id': annee_trimestre.id_trimestre,
            'trimestre': annee_trimestre.trimestre.trimestre,
            'etat': annee_trimestre.etat_trimestre,
            'debut': annee_trimestre.debut,
            'fin': annee_trimestre.fin
        },
        'variables': variables_data
    })


def get_eleves_classe(request):
    id_campus = request.GET.get('id_campus')
    id_cycle_actif = request.GET.get('id_cycle')
    id_classe_active = request.GET.get('id_classe_active')
    id_annee = request.GET.get('id_annee')

    # 1️⃣ récupérer la classe active
    classe_active = Classe_active.objects.filter(
        id_classe_active=id_classe_active,
        id_campus_id=id_campus,
        id_annee_id=id_annee,
        is_active=True
    ).first()

    if not classe_active:
        return JsonResponse([], safe=False)

    # 2️⃣ récupérer les élèves inscrits dans cette classe
    eleves = Eleve_inscription.objects.filter(
        id_campus_id=id_campus,
        id_annee_id=id_annee,
        id_classe_id=id_classe_active,   # ✅ correspondance réelle
        id_classe_cycle_id=id_cycle_actif,
        status=1
    ).select_related('id_eleve').distinct()

    data = [{
        "id_eleve": e.id_eleve.id_eleve,
        "nom": e.id_eleve.nom,
        "prenom": e.id_eleve.prenom
    } for e in eleves]

    return JsonResponse(data, safe=False)


def historique_financier(request):
    def clean(v): 
        return v if v not in ("", None) else None

    id_annee = clean(request.GET.get('annee'))
    id_classe = clean(request.GET.get('classe'))
    id_eleve = clean(request.GET.get('eleve'))
    id_trimestre = clean(request.GET.get('trimestre'))
    id_compte = clean(request.GET.get('compte'))

    # Récupérer d'abord les IDs des VariablePrix correspondant aux critères
    variable_prix_ids = []
    
    if id_trimestre or id_annee or id_classe:
        # Construire la requête sur VariablePrix
        variable_prix_qs = VariablePrix.objects.all()
        
        if id_annee:
            variable_prix_qs = variable_prix_qs.filter(id_annee_id=id_annee)
        if id_classe:
            variable_prix_qs = variable_prix_qs.filter(id_classe_active_id=id_classe)
        if id_trimestre:
            variable_prix_qs = variable_prix_qs.filter(id_annee_trimestre_id=id_trimestre)
        
        # Récupérer les IDs des variables concernées
        variable_prix_ids = list(variable_prix_qs.values_list('id_variable_id', flat=True).distinct())
    
    # Maintenant, requête sur les paiements
    paiements_qs = Paiement.objects.select_related(
        'id_eleve',
        'id_variable',
        'id_compte',
        'id_compte__id_banque'
    ).order_by(
        'id_eleve__nom',
        'id_variable__variable',
        'date_paie'
    )

    if id_annee:
        paiements_qs = paiements_qs.filter(id_annee=id_annee)
    
    if id_classe:
        paiements_qs = paiements_qs.filter(id_classe_active=id_classe)
    
    if id_eleve:
        paiements_qs = paiements_qs.filter(id_eleve=id_eleve)
    
    if id_trimestre and variable_prix_ids:
        # Filtrer par les variables qui sont dans VariablePrix pour ce trimestre
        paiements_qs = paiements_qs.filter(id_variable_id__in=variable_prix_ids)
    elif id_trimestre and not variable_prix_ids:
        # Si aucun VariablePrix trouvé pour ce trimestre, retourner vide
        return JsonResponse({
            "success": True,
            "rapport": [],
            "total_general": 0,
            "message": "Aucune variable trouvée pour ce trimestre"
        })
    
    if id_compte:
        paiements_qs = paiements_qs.filter(id_compte=id_compte)

    rapport = []
    total_general = 0

    for p in paiements_qs:
        total_general += p.montant

        rapport.append({
            "eleve": f"{p.id_eleve.nom} {p.id_eleve.prenom}",
            "variable": p.id_variable.variable,
            "montant": p.montant,
            "date_paie": p.date_paie.strftime('%d/%m/%Y'),
            "banque_compte": f"{p.id_compte.id_banque.banque} - {p.id_compte.compte}"
        })

    return JsonResponse({
        "success": True,
        "rapport": rapport,
        "total_general": total_general
    })

def eleves_en_dette(request):

    id_annee = request.GET.get('annee')
    id_classe = request.GET.get('classe')
    id_trimestre = request.GET.get('trimestre')

    if not id_annee or not id_classe:
        return JsonResponse({
            "success": False,
            "message": "Année et classe obligatoires"
        })

    # 🔹 1. Récupérer les INSCRIPTIONS (BASE)
    inscriptions = Eleve_inscription.objects.select_related(
        'id_eleve',
        'id_classe',
        'id_classe__classe_id'
    ).filter(
        id_annee=id_annee,
        id_classe=id_classe,
        status=True
    )

    if id_trimestre:
        inscriptions = inscriptions.filter(id_trimestre=id_trimestre)

    rapport = []

    for ins in inscriptions:
        eleve = ins.id_eleve

        # Nom classe + groupe
        nom_classe = ins.id_classe.classe_id.classe
        groupe = ins.id_classe.groupe or ""
        classe_info = f"{nom_classe} {groupe}".strip()

        # 🔹 Paiements de l'élève
        paiements = Paiement.objects.filter(
            id_eleve=eleve,
            id_classe_active=id_classe,
            id_annee=id_annee
        )

        # 🔹 Variables attendues pour la classe
        variables = VariablePrix.objects.filter(
            id_classe_active=id_classe,
            id_annee=id_annee,
            id_variable__estObligatoire=True
        ).select_related('id_variable')

        details = []
        total_dette = 0

        for vp in variables:
            variable = vp.id_variable

            montant_a_payer = vp.prix

            montant_paye = paiements.filter(
                id_variable=variable
            ).aggregate(total=Sum('montant'))['total'] or 0

            reste = max(montant_a_payer - montant_paye, 0)

            # 🔹 pénalité (simple)
            penalite = 0
            if reste > 0:
                penalite = PenaliteConfig.objects.filter(
                    id_variable=variable,
                    id_classe_active=id_classe,
                    actif=True
                ).aggregate(total=Sum('valeur'))['total'] or 0

            total_variable = reste + penalite

            if total_variable > 0:
                details.append({
                    "variable": variable.variable,
                    "montant_a_payer": montant_a_payer,
                    "montant_paye": montant_paye,
                    "reste": reste,
                    "penalite": penalite,
                    "total": total_variable
                })

                total_dette += total_variable

        if total_dette > 0:
            rapport.append({
                "id_eleve": eleve.id_eleve,
                "eleve": f"{eleve.nom} {eleve.prenom}",
                "classe": classe_info,
                "total_dette": total_dette,
                "details": details
            })

    return JsonResponse({
        "success": True,
        "rapport": rapport
    })


def get_penalites(request):
    penalites = PenaliteConfig.objects.select_related(
        'id_annee',
        'id_campus',
        'id_cycle_actif',
        'id_classe_active',
        'id_variable',
        'id_annee_trimestre'
    ).order_by('-id_penalite_regle')

    data = []

    for p in penalites:
        data.append({
            'id_penalite_regle': p.id_penalite_regle,
            'annee': str(p.id_annee) if p.id_annee else '',
            'campus': str(p.id_campus) if p.id_campus else '',
            'cycle': str(p.id_cycle_actif) if p.id_cycle_actif else '',
            'classe': str(p.id_classe_active) if p.id_classe_active else '',
            'variable': str(p.id_variable) if p.id_variable else '',
            'trimestre': str(p.id_annee_trimestre) if p.id_annee_trimestre else '',
            'type_penalite': p.type_penalite,
            'valeur': p.valeur,
            'plafond': p.plafond if p.plafond else '',
            'actif': p.actif
        })

    return JsonResponse({'success': True, 'penalites': data})

@csrf_exempt
def toggle_penalite_actif(request):
    if request.method == 'POST':
        penalite_id = request.POST.get('id_penalite')
        actif = request.POST.get('actif') == 'true'

        try:
            penalite = PenaliteConfig.objects.get(id_penalite_regle=penalite_id)
            penalite.actif = actif
            penalite.save()
            return JsonResponse({'success': True})
        except PenaliteConfig.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Pénalité non trouvée'})
    return JsonResponse({'success': False, 'error': 'Méthode non autorisée'})

def get_classes_with_penalite(request):
    """Retourne uniquement les classes actives qui ont au moins une PenaliteConfig."""
    id_annee = request.GET.get('id_annee')
    if not id_annee:
        return JsonResponse({'success': False, 'error': 'Année requise'}, status=400)

    try:
        # Classes qui ont une pénalité configurée
        classe_ids = PenaliteConfig.objects.filter(
            id_annee_id=id_annee,
            actif=True,
            id_classe_active__isnull=False
        ).values_list('id_classe_active_id', flat=True).distinct()

        classes = Classe_active.objects.filter(
            id_classe_active__in=classe_ids
        ).select_related('id_campus', 'cycle_id__cycle_id', 'classe_id')

        data = [{
            'id_classe_active': c.id_classe_active,
            'campus_nom': c.id_campus.campus,
            'cycle_nom': c.cycle_id.cycle_id.cycle,
            'classe_nom': c.classe_id.classe,
            'groupe': c.groupe or '',
            'id_campus': c.id_campus.id_campus,
            'id_cycle': c.cycle_id.id_cycle_actif,
        } for c in classes]

        return JsonResponse({'success': True, 'classes': data})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


def get_trimestres_with_penalite(request):
    """Retourne uniquement les trimestres ayant une PenaliteConfig pour la classe sélectionnée."""
    id_annee = request.GET.get('id_annee')
    id_classe_active = request.GET.get('id_classe_active')

    if not all([id_annee, id_classe_active]):
        return JsonResponse({'success': False, 'error': 'Paramètres manquants'}, status=400)

    try:
        trimestre_ids = PenaliteConfig.objects.filter(
            id_annee_id=id_annee,
            id_classe_active_id=id_classe_active,
            actif=True,
            id_annee_trimestre__isnull=False
        ).values_list('id_annee_trimestre_id', flat=True).distinct()

        trimestres = Annee_trimestre.objects.filter(
            id_trimestre__in=trimestre_ids
        ).select_related('trimestre')

        data = [{
            'id': t.id_trimestre,
            'nom': t.trimestre.trimestre,
            'etat': t.etat_trimestre,
        } for t in trimestres]

        return JsonResponse({'success': True, 'trimestres': data})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


def get_variables_with_penalite(request):
    """Retourne uniquement les variables ayant une PenaliteConfig pour les filtres sélectionnés."""
    id_annee = request.GET.get('id_annee')
    id_classe_active = request.GET.get('id_classe_active')
    id_trimestre = request.GET.get('id_trimestre')

    if not id_annee:
        return JsonResponse({'success': False, 'error': 'Année requise'}, status=400)

    try:
        filtre = Q(id_annee_id=id_annee, actif=True, id_variable__isnull=False)
        if id_classe_active:
            filtre &= Q(id_classe_active_id=id_classe_active)
        if id_trimestre:
            filtre &= Q(id_annee_trimestre_id=id_trimestre)

        variable_ids = PenaliteConfig.objects.filter(filtre).values_list('id_variable_id', flat=True).distinct()

        variables = Variable.objects.filter(id_variable__in=variable_ids)

        data = [{
            'id_variable': v.id_variable,
            'variable': v.variable,
        } for v in variables]

        return JsonResponse({'success': True, 'variables': data})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


def eleves_en_penalite(request):

    id_campus   = request.GET.get('id_campus')
    id_cycle    = request.GET.get('id_cycle')
    id_classe   = request.GET.get('classe')
    annee       = request.GET.get('annee')
    variable_id = request.GET.get('variable')
    trimestre   = request.GET.get('trimestre')

    if not annee:
        return JsonResponse({'success': False, 'error': 'Veuillez sélectionner une année.'})

    # ================================
    # FILTRE INSCRIPTION ELEVE
    # ================================
    filtre_eleve = Q(id_annee_id=annee, status=True)

    if id_campus:
        filtre_eleve &= Q(id_campus_id=id_campus)

    if id_cycle:
        filtre_eleve &= Q(id_classe_cycle_id=id_cycle)

    if id_classe:
        filtre_eleve &= Q(id_classe_id=id_classe)

    if trimestre:
        filtre_eleve &= Q(id_trimestre_id=trimestre)

    # IMPORTANT : select_related pour performance
    inscriptions = Eleve_inscription.objects.filter(
        filtre_eleve
    ).select_related('id_eleve')

    # ================================
    # FILTRE CONFIG PENALITE
    # ================================
    filtre_config = Q(id_annee_id=annee, actif=True)

    if variable_id:
        filtre_config &= Q(id_variable_id=variable_id)

    if trimestre:
        filtre_config &= Q(id_annee_trimestre_id=trimestre)

    if id_campus:
        filtre_config &= Q(id_campus_id=id_campus)

    if id_cycle:
        filtre_config &= Q(id_cycle_actif_id=id_cycle)

    configs = PenaliteConfig.objects.filter(filtre_config)

    resultats = []

    # ================================
    # LOGIQUE PENALITE
    # ================================
    for config in configs:

        db_obj = VariableDatebutoire.objects.filter(
            id_variable=config.id_variable,
            id_annee=config.id_annee,
            id_campus=config.id_campus,
            id_cycle_actif=config.id_cycle_actif,
            id_variable__estObligatoire=True
        ).first()

        if not db_obj:
            continue

        eleves_concernes = inscriptions.filter(
            id_campus=config.id_campus,
            id_classe_cycle=config.id_cycle_actif
        )

        for insc in eleves_concernes:

            eleve = insc.id_eleve   # ✅ LE BON OBJET ELEVE

            # ================= DEROGATION
            derog = VariableDerogation.objects.filter(
                id_eleve=eleve,
                id_variable=config.id_variable,
                id_annee_id=annee,
                id_variable__estObligatoire=True
            ).first()

            date_limite = derog.date_derogation if derog else db_obj.date_butoire

            # ================= PAIEMENT
            paiement = Paiement.objects.filter(
                id_eleve=eleve,
                id_variable=config.id_variable,
                id_variable__estObligatoire=True,
                id_annee_id=annee,
                status=True
            ).order_by('date_paie').first()

            if paiement and paiement.date_paie and paiement.date_paie <= date_limite:
                continue

            # ================= PRIX
            prix_obj = VariablePrix.objects.filter(
                id_variable=config.id_variable,
                id_annee_id=annee,
                id_campus=config.id_campus,
                id_cycle_actif=config.id_cycle_actif,
                id_variable__estObligatoire=True
            ).first()

            montant_base = prix_obj.prix if prix_obj else 0

            reduc = Eleve_reduction_prix.objects.filter(
                id_eleve=eleve,
                id_variable=config.id_variable
            ).first()

            if reduc:
                montant_base -= (montant_base * reduc.pourcentage / 100)

            # ================= PENALITE
            if config.type_penalite == 'FORFAIT':
                montant_p = config.valeur
            else:
                montant_p = (montant_base * config.valeur / 100)
                if config.plafond:
                    montant_p = min(montant_p, config.plafond)

            resultats.append({
                'eleve': f"{eleve.nom} {eleve.prenom}",
                'variable': config.id_variable.variable,
                'date_limite': date_limite.strftime('%d/%m/%Y'),
                'montant_penalite': round(montant_p, 2)
            })

    return JsonResponse({'success': True, 'eleves': resultats})


def suivi_reduction_derogation_data(request):

    type_filter = request.GET.get('type')
    annee = request.GET.get('annee')
    classe = request.GET.get('classe')
    eleve = request.GET.get('eleve')
    variable = request.GET.get('variable')

    rows = []

    # ========================
    # REDUCTIONS
    # ========================
    if type_filter in ['reduction', 'all', '']:

        reductions = Eleve_reduction_prix.objects.filter(id_annee=annee)

        if classe:
            reductions = reductions.filter(id_classe_active=classe)
        if eleve:
            reductions = reductions.filter(id_eleve=eleve)
        if variable:
            reductions = reductions.filter(id_variable=variable)

        for r in reductions:
            rows.append({
                'id': r.id_reduction_prix,
                'type': 'reduction',
                'eleve': f'{r.id_eleve.nom} {r.id_eleve.prenom}',
                'classe': str(r.id_classe_active.classe_id),
                'variable': r.id_variable.variable,
                'statut': f'{r.pourcentage}',
            })

    # ========================
    # DEROGATIONS
    # ========================
    if type_filter in ['derogation', 'all', '']:

        derogations = VariableDerogation.objects.filter(id_annee=annee)

        if classe:
            derogations = derogations.filter(id_classe_active=classe)
        if eleve:
            derogations = derogations.filter(id_eleve=eleve)
        if variable:
            derogations = derogations.filter(id_variable=variable)

        for d in derogations:
            rows.append({
                'id': d.id_derogation,
                'type': 'derogation',
                'eleve': f'{d.id_eleve.nom} {d.id_eleve.prenom}',
                'classe': str(d.id_classe_active.classe_id),
                'variable': d.id_variable.variable,
                'statut': str(d.date_derogation),
            })

    return JsonResponse({'success': True, 'rows': rows})

def get_dates_butoire(request):
    try:
        dates = VariableDatebutoire.objects.select_related(
            "id_variable",
            "id_classe_active"
        )

        rows = []
        for d in dates:
            rows.append({
                "id": d.id_datebutoire,
                "classe": str(d.id_classe_active),
                "trimestre": str(d.trimestre) if hasattr(d, "trimestre") else None,
                "variable": d.id_variable.variable,
                "date_butoire": d.date_butoire.strftime("%Y-%m-%d")
            })

        return JsonResponse({
            "success": True,
            "rows": rows
        })

    except Exception as e:
        return JsonResponse({
            "success": False,
            "error": str(e)
        })



def situation_journaliere_data(request):

    date_debut = request.GET.get('date_debut')
    date_fin = request.GET.get('date_fin')

    # 🔥 IMPORTANT : convertir les chaines vides en None
    if not date_debut:
        date_debut = None
    if not date_fin:
        date_fin = None

    paiements = Paiement.objects.filter(
        status=True,
        is_rejected=False
    )

    # =========================
    # FILTRE DATE
    # =========================
    if date_debut and date_fin:
        paiements = paiements.filter(date_saisie__range=[date_debut, date_fin])
    else:
        # ✅ PAR DEFAUT : AUJOURD'HUI
        today = date.today()
        paiements = paiements.filter(date_saisie=today)

    total = paiements.aggregate(total=Sum('montant'))['total'] or 0

    rows = []
    # print("DEBUG paiements trouvés:", paiements)

    for p in paiements.select_related('id_eleve','id_variable'):
        rows.append({
            "eleve": str(p.id_eleve),
            "variable": str(p.id_variable),
            "montant": p.montant,
            "compte": str(p.id_compte) if p.id_compte else None,
            "date_paie": p.date_paie.strftime("%Y-%m-%d"),
        })

    return JsonResponse({
        "success": True,
        "rows": rows,
        "total": total
    })



def rapport_paiements(request):
    id_annee = request.GET.get("id_annee")
    id_campus = request.GET.get("id_campus")
    id_cycle = request.GET.get("id_cycle")
    id_classe_active = request.GET.get("id_classe_active")
    date_debut = request.GET.get("date_debut")
    date_fin = request.GET.get("date_fin")
    type_fichier = request.GET.get("type_fichier")

    if not id_annee or not id_classe_active:
        return HttpResponse("Paramètres manquants")

    # Conversion des dates
    if date_debut:
        try:
            date_debut = datetime.strptime(date_debut, '%Y-%m-%d').date()
        except ValueError:
            date_debut = None
    if date_fin:
        try:
            date_fin = datetime.strptime(date_fin, '%Y-%m-%d').date()
        except ValueError:
            date_fin = None

    # ============================
    # CLASSE
    # ============================
    try:
        classe = Classe_active.objects.select_related(
            "id_campus",
            "cycle_id"
        ).get(pk=id_classe_active)
    except Classe_active.DoesNotExist:
        return HttpResponse("Classe non trouvée")

    # ============================
    # INSCRIPTIONS
    # ============================
    inscriptions = Eleve_inscription.objects.select_related(
        "id_eleve"
    ).filter(
        id_classe=id_classe_active,
        id_annee=id_annee
    )

    # ============================
    # PRIX PAR VARIABLE POUR LA CLASSE
    # ============================
    variables_prix = VariablePrix.objects.filter(
        id_annee=id_annee,
        id_classe_active=id_classe_active
    ).select_related('id_variable', 'id_annee_trimestre')

    # Dictionnaire pour stocker les prix par variable
    prix_par_variable = {}
    for vp in variables_prix:
        # Nom de variable plus court pour l'affichage
        nom_variable = vp.id_variable.variable if hasattr(vp.id_variable, 'variable') else f"Var-{vp.id_variable.id_variable}"
        # Raccourcir les noms longs
        if "Minerval" in nom_variable:
            nom_variable = "Minerval"
        elif "Inscription" in nom_variable:
            nom_variable = "Inscription"
        elif "Evaluation" in nom_variable or "évaluation" in nom_variable:
            nom_variable = "Éval"
            
        prix_par_variable[vp.id_variable.id_variable] = {
            'prix': vp.prix,
            'trimestre': vp.id_annee_trimestre,
            'variable_obj': vp.id_variable,
            'nom_variable': nom_variable
        }

    # ============================
    # DATES BUTOIRES
    # ============================
    dates_butoires = VariableDatebutoire.objects.filter(
        id_annee=id_annee,
        id_classe_active=id_classe_active
    ).select_related('id_variable')

    # Dictionnaire des dates butoires par variable
    date_butoire_par_variable = {}
    for db in dates_butoires:
        date_butoire_par_variable[db.id_variable.id_variable] = db.date_butoire

    # ============================
    # CONFIGURATION DES PÉNALITÉS
    # ============================
    penalites_config = PenaliteConfig.objects.filter(
        id_annee=id_annee,
        actif=True
    ).filter(
        Q(id_classe_active=id_classe_active) | 
        Q(id_classe_active__isnull=True)
    ).filter(
        Q(id_campus=classe.id_campus) | 
        Q(id_campus__isnull=True)
    )

    # ============================
    # TABLEAUX DEMANDÉS
    # ============================
    tableau_complet = []      # Élèves ayant tout payé dans la période
    tableau_reduction = []    # Élèves avec réductions
    tableau_derogation = []   # Élèves avec dérogations
    tableau_penalite = []     # Élèves avec pénalités
    tableau_dette = []        # Élèves ayant une dette

    # ============================
    # TRAITEMENT PAR ÉLÈVE
    # ============================
    for ins in inscriptions:
        eleve = ins.id_eleve

        # Nom complet de l'élève
        nom_complet = ""
        if hasattr(eleve, 'nom') and hasattr(eleve, 'prenom'):
            nom_complet = f"{eleve.nom} {eleve.prenom}".strip()
        elif hasattr(eleve, 'nom'):
            nom_complet = eleve.nom
        elif hasattr(eleve, 'prenom'):
            nom_complet = eleve.prenom
        else:
            nom_complet = str(eleve)

        # ======================
        # VÉRIFIER LES RÉDUCTIONS
        # ======================
        reductions = Eleve_reduction_prix.objects.filter(
            id_eleve=eleve,
            id_annee=id_annee,
            id_classe_active=id_classe_active
        ).select_related('id_variable')

        reduction_par_variable = {}
        for red in reductions:
            reduction_par_variable[red.id_variable.id_variable] = red.pourcentage

        # ======================
        # VÉRIFIER LES DÉROGATIONS
        # ======================
        derogations = VariableDerogation.objects.filter(
            id_eleve=eleve,
            id_annee=id_annee,
            id_classe_active=id_classe_active
        ).select_related('id_variable')

        derogation_par_variable = {}
        for der in derogations:
            derogation_par_variable[der.id_variable.id_variable] = der.date_derogation

        # ======================
        # CALCULER CE QUI ÉTAIT DÛ PENDANT LA PÉRIODE
        # ======================
        montant_du_periode = 0
        details_du = []
        variables_dues = []

        for var_id, var_info in prix_par_variable.items():
            prix_variable = var_info['prix']
            nom_variable = var_info['nom_variable']
            
            # Vérifier si cette variable était due dans la période
            est_dans_periode = False
            
            if var_id in date_butoire_par_variable:
                date_butoire = date_butoire_par_variable[var_id]
                if date_debut and date_fin:
                    if date_debut <= date_butoire <= date_fin:
                        est_dans_periode = True
                elif date_debut and not date_fin:
                    if date_butoire >= date_debut:
                        est_dans_periode = True
                elif not date_debut and date_fin:
                    if date_butoire <= date_fin:
                        est_dans_periode = True
                else:
                    est_dans_periode = True
            else:
                # Si pas de date butoire, on considère par défaut
                est_dans_periode = True

            if est_dans_periode:
                # Appliquer la réduction si elle existe
                if var_id in reduction_par_variable:
                    pourcentage = reduction_par_variable[var_id]
                    reduction = (prix_variable * pourcentage) / 100
                    prix_apres_reduction = prix_variable - reduction
                else:
                    prix_apres_reduction = prix_variable
                    reduction = 0

                montant_du_periode += prix_apres_reduction
                variables_dues.append(var_id)
                
                details_du.append({
                    'variable': nom_variable,
                    'prix_initial': prix_variable,
                    'reduction': reduction,
                    'prix_final': prix_apres_reduction,
                    'a_reduction': var_id in reduction_par_variable,
                    'a_derogation': var_id in derogation_par_variable,
                    'date_butoire': date_butoire_par_variable.get(var_id, None)
                })

        # ======================
        # CALCULER LES PAIEMENTS PENDANT LA PÉRIODE
        # ======================
        paiements = Paiement.objects.filter(
            id_eleve=eleve,
            id_annee=id_annee,
            id_classe_active=id_classe_active,
            status=True,
            is_rejected=False
        )

        if date_debut and date_fin:
            paiements = paiements.filter(
                date_saisie__range=[date_debut, date_fin]
            )
        elif date_debut and not date_fin:
            paiements = paiements.filter(date_saisie__gte=date_debut)
        elif not date_debut and date_fin:
            paiements = paiements.filter(date_saisie__lte=date_fin)

        # Paiements par variable
        paiements_par_variable = {}
        for paiement in paiements:
            var_id = paiement.id_variable.id_variable
            if var_id not in paiements_par_variable:
                paiements_par_variable[var_id] = 0
            paiements_par_variable[var_id] += paiement.montant

        total_paye_periode = sum(paiements_par_variable.values())
        reste_a_payer = montant_du_periode - total_paye_periode

        # ======================
        # VÉRIFIER LES PÉNALITÉS
        # ======================
        penalite_calculee = 0
        a_penalite = False
        details_penalite = []

        for var_id, var_info in prix_par_variable.items():
            if var_id in variables_dues and var_id in date_butoire_par_variable:
                date_butoire = date_butoire_par_variable[var_id]
                nom_variable = var_info['nom_variable']
                
                # Vérifier si des paiements ont été faits après la date butoire
                paiements_tardifs = paiements.filter(
                    id_variable=var_id,
                    date_saisie__gt=date_butoire
                )
                
                if paiements_tardifs.exists():
                    a_penalite = True
                    montant_tardif = sum(paiements_tardifs.values_list('montant', flat=True))
                    
                    # Chercher la configuration de pénalité appropriée
                    config_penalite = penalites_config.filter(
                        Q(id_variable=var_id) | Q(id_variable__isnull=True)
                    ).first()
                    
                    if config_penalite:
                        if config_penalite.type_penalite == 'POURCENTAGE':
                            penalite = montant_tardif * config_penalite.valeur / 100
                        else:  # FORFAIT
                            penalite = config_penalite.valeur
                        
                        # Appliquer le plafond si existant
                        if config_penalite.plafond and penalite > config_penalite.plafond:
                            penalite = config_penalite.plafond
                    else:
                        penalite = montant_tardif * 0.05  # 5% par défaut
                    
                    penalite_calculee += penalite
                    details_penalite.append(f"{nom_variable}: {penalite:,.0f}")

        # ======================
        # CONSTRUCTION DES LIGNES POUR CHAQUE TABLEAU
        # ======================

        # TABLEAU DES PAIEMENTS COMPLETS
        if montant_du_periode > 0 and total_paye_periode >= montant_du_periode:
            row_complet = [
                nom_complet,
                f"{montant_du_periode:,.0f}",
                f"{total_paye_periode:,.0f}",
                "COMPLET"
            ]
            tableau_complet.append(row_complet)

        # TABLEAU DES RÉDUCTIONS
        if reductions.exists():
            details_reduction = []
            for detail in details_du:
                if detail['a_reduction']:
                    pourcentage = reduction_par_variable.get(
                        [k for k, v in prix_par_variable.items() if v['nom_variable'] == detail['variable']][0], 
                        0
                    )
                    details_reduction.append(
                        f"{detail['variable']}: {pourcentage}%"
                    )
            
            if details_reduction:
                row_reduction = [
                    nom_complet,
                    f"{montant_du_periode:,.0f}",
                    " / ".join(details_reduction),
                    f"{total_paye_periode:,.0f}",
                    f"{reste_a_payer:,.0f}" if reste_a_payer > 0 else "0"
                ]
                tableau_reduction.append(row_reduction)

        # TABLEAU DES DÉROGATIONS
        if derogations.exists():
            details_derogation = []
            for der in derogations:
                var_nom = prix_par_variable.get(der.id_variable.id_variable, {}).get('nom_variable', '')
                if var_nom:
                    details_derogation.append(
                        f"{var_nom}: {der.date_derogation.strftime('%d/%m')}"
                    )
            
            if details_derogation:
                row_derogation = [
                    nom_complet,
                    f"{montant_du_periode:,.0f}",
                    " / ".join(details_derogation),
                    f"{total_paye_periode:,.0f}",
                    f"{reste_a_payer:,.0f}" if reste_a_payer > 0 else "0"
                ]
                tableau_derogation.append(row_derogation)

        # TABLEAU DES PÉNALITÉS
        if a_penalite:
            # Formater les détails des pénalités avec des sauts de ligne
            if len(details_penalite) > 1:
                penalite_display = "\n".join(details_penalite)
            else:
                penalite_display = details_penalite[0] if details_penalite else f"{penalite_calculee:,.0f}"
            
            row_penalite = [
                nom_complet,
                f"{montant_du_periode:,.0f}",
                penalite_display,
                f"{total_paye_periode:,.0f}",
                f"{reste_a_payer:,.0f}" if reste_a_payer > 0 else "0"
            ]
            tableau_penalite.append(row_penalite)

        # TABLEAU DES DETTES
        if reste_a_payer > 0:
            row_dette = [
                nom_complet,
                f"{montant_du_periode:,.0f}",
                f"{total_paye_periode:,.0f}",
                f"{reste_a_payer:,.0f}"
            ]
            tableau_dette.append(row_dette)

    # ============================
    # EN-TÊTES DES TABLEAUX (cohérents et sans erreurs)
    # ============================
    headers_complet = [
        "Élève",
        "Dû",
        "Payé",
        "Statut"
    ]

    headers_reduction = [
        "Élève",
        "Dû",
        "Réductions",
        "Payé",
        "Reste"
    ]

    headers_derogation = [
        "Élève",
        "Dû",
        "Dérogations",
        "Payé",
        "Reste"
    ]

    headers_penalite = [
        "Élève",
        "Dû",
        "Pénalités",
        "Payé",
        "Reste"
    ]

    headers_dette = [
        "Élève",
        "Dû",
        "Payé",
        "Dette"
    ]

    # ============================
    # EXPORT
    # ============================
    if type_fichier == "excel":
        return export_excel_multi(
            tableau_complet,
            tableau_reduction,
            tableau_derogation,
            tableau_penalite,
            tableau_dette,
            headers_complet,
            headers_reduction,
            headers_derogation,
            headers_penalite,
            headers_dette,
            date_debut,
            date_fin,
            classe
        )

    if type_fichier == "pdf":
        return export_pdf_multi(
            tableau_complet,
            tableau_reduction,
            tableau_derogation,
            tableau_penalite,
            tableau_dette,
            headers_complet,
            headers_reduction,
            headers_derogation,
            headers_penalite,
            headers_dette,
            date_debut,
            date_fin,
            classe
        )

    return HttpResponse("Type non supporté")


def add_section_excel(ws, title, data, headers, start_row):
    """Ajoute une section au fichier Excel avec un style amélioré"""
    
    # Styles
    title_font = Font(bold=True, size=12, color="FFFFFF")
    title_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    current_row = start_row
    
    # Titre de section
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(headers))
    cell = ws.cell(row=current_row, column=1, value=title)
    cell.font = title_font
    cell.fill = title_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    current_row += 1
    
    # Nombre d'élèves
    ws.cell(row=current_row, column=1, value=f"Nombre : {len(data)}")
    current_row += 1
    
    # En-têtes
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=current_row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    current_row += 1
    
    # Données
    for row_data in data:
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=current_row, column=col, value=value)
            cell.border = border
            if col == 1:  # Colonne Élève
                cell.alignment = Alignment(horizontal='left', vertical='center')
            else:  # Colonnes de montants
                cell.alignment = Alignment(horizontal='right', vertical='center')
            # Activer le retour à la ligne pour les cellules avec plusieurs lignes
            if col == 3 and isinstance(value, str) and "\n" in value:
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        current_row += 1
    
    return current_row


def export_excel_multi(tableau_complet, tableau_reduction, tableau_derogation, 
                      tableau_penalite, tableau_dette, headers_complet, 
                      headers_reduction, headers_derogation, headers_penalite, 
                      headers_dette, date_debut, date_fin, classe):

    wb = Workbook()
    ws = wb.active
    ws.title = "Rapport Paiements"

    # Style pour le titre principal
    title_font = Font(bold=True, size=16, color="FFFFFF")
    title_fill = PatternFill(start_color="2F528F", end_color="2F528F", fill_type="solid")
    
    current_row = 1
    
    # Informations de période et classe
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=5)
    cell = ws.cell(row=current_row, column=1)
    if date_debut and date_fin:
        cell.value = f"RAPPORT PAIEMENTS - {classe} - Du {date_debut} au {date_fin}"
    elif date_debut:
        cell.value = f"RAPPORT PAIEMENTS - {classe} - A partir du {date_debut}"
    elif date_fin:
        cell.value = f"RAPPORT PAIEMENTS - {classe} - Jusqu'au {date_fin}"
    else:
        cell.value = f"RAPPORT PAIEMENTS - {classe}"
    cell.font = title_font
    cell.fill = title_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    
    current_row += 1
    ws.cell(row=current_row, column=1, value=f"Edite le {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    current_row += 2

    # ============================================
    # SECTIONS - Uniquement celles avec des données
    # ============================================
    sections = []
    
    # N'ajouter que les sections qui ont des données
    if tableau_complet:
        sections.append((tableau_complet, "1. PAIEMENTS COMPLETS", headers_complet))
    if tableau_reduction:
        sections.append((tableau_reduction, "2. RÉDUCTIONS", headers_reduction))
    if tableau_derogation:
        sections.append((tableau_derogation, "3. DÉROGATIONS", headers_derogation))
    if tableau_penalite:
        sections.append((tableau_penalite, "4. PÉNALITÉS", headers_penalite))
    if tableau_dette:
        sections.append((tableau_dette, "5. DETTES", headers_dette))
    
    # Renumérotation continue
    for idx, (data, title, headers) in enumerate(sections, 1):
        # Remplacer le premier chiffre par le nouvel index
        nouveau_titre = title.replace(title[0], str(idx), 1)
        
        # Titre de section
        ws.cell(row=current_row, column=1, value=nouveau_titre).font = Font(bold=True, size=11, color="FFFFFF")
        ws.cell(row=current_row, column=1).fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        current_row += 1
        
        # En-têtes
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                                top=Side(style='thin'), bottom=Side(style='thin'))
            cell.alignment = Alignment(horizontal='center', vertical='center')
        current_row += 1
        
        # Données
        for row_data in data:
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=current_row, column=col, value=value)
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                                    top=Side(style='thin'), bottom=Side(style='thin'))
                if col == 1:
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                else:
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                # Activer le retour à la ligne pour les cellules avec sauts de ligne
                if col == 3 and isinstance(value, str) and "\n" in value:
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            current_row += 1
        
        current_row += 1

    # ============================================
    # RÉCAPITULATIF
    # ============================================
    total_eleves_classe = Eleve_inscription.objects.filter(id_classe=classe.id_classe_active).count()
    
    current_row += 1
    ws.cell(row=current_row, column=1, value="RÉCAPITULATIF").font = Font(bold=True, size=11)
    current_row += 1
    
    recap_data = [
        ["Catégorie", "Nombre d'élèves", "Pourcentage"],
        ["Complets", len(tableau_complet), f"{len(tableau_complet)/total_eleves_classe*100:.1f}%" if total_eleves_classe > 0 else "0%"],
        ["Réductions", len(tableau_reduction), f"{len(tableau_reduction)/total_eleves_classe*100:.1f}%" if total_eleves_classe > 0 else "0%"],
        ["Dérogations", len(tableau_derogation), f"{len(tableau_derogation)/total_eleves_classe*100:.1f}%" if total_eleves_classe > 0 else "0%"],
        ["Pénalités", len(tableau_penalite), f"{len(tableau_penalite)/total_eleves_classe*100:.1f}%" if total_eleves_classe > 0 else "0%"],
        ["Endettés", len(tableau_dette), f"{len(tableau_dette)/total_eleves_classe*100:.1f}%" if total_eleves_classe > 0 else "0%"],
        ["TOTAL", total_eleves_classe, "100%"]
    ]
    
    # En-têtes du récapitulatif
    for col, value in enumerate(recap_data[0], 1):
        cell = ws.cell(row=current_row, column=col, value=value)
        cell.font = Font(bold=True)
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                            top=Side(style='thin'), bottom=Side(style='thin'))
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
    current_row += 1
    
    # Données du récapitulatif
    for row_data in recap_data[1:]:
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=current_row, column=col, value=value)
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                                top=Side(style='thin'), bottom=Side(style='thin'))
            if row_data[0] == "TOTAL":
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
            if col == 1:
                cell.alignment = Alignment(horizontal='left')
            else:
                cell.alignment = Alignment(horizontal='center')
        current_row += 1

    # ============================================
    # AJUSTEMENT DES LARGEURS DE COLONNES
    # ============================================
    for col in range(1, ws.max_column + 1):
        max_length = 0
        column_letter = get_column_letter(col)
        for row in range(1, ws.max_row + 1):
            cell = ws.cell(row=row, column=col)
            try:
                if cell.value:
                    if isinstance(cell.value, str) and "\n" in cell.value:
                        lines = cell.value.split("\n")
                        max_line_length = max(len(line) for line in lines)
                        max_length = max(max_length, max_line_length)
                    else:
                        max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        
        # Largeurs adaptatives selon le type de colonne
        if col == 1:  # Colonne Élève
            adjusted_width = min(max_length + 2, 35)
        elif col == 3 and any([tableau_reduction, tableau_derogation, tableau_penalite]):  # Colonne des détails
            adjusted_width = min(max_length + 4, 50)
        elif col <= 3:  # Autres colonnes
            adjusted_width = min(max_length + 2, 20)
        else:
            adjusted_width = min(max_length + 2, 15)
            
        ws.column_dimensions[column_letter].width = adjusted_width

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="rapport_paiements.xlsx"'
    wb.save(response)

    return response


def calculate_dynamic_widths(table_data, headers):
    """
    Calcule les largeurs de colonnes de façon dynamique en fonction du contenu
    Retourne une liste de largeurs en cm
    """
    col_widths = []
    
    for col_idx in range(len(headers)):
        # Commencer par la largeur de l'en-tête
        max_chars = len(headers[col_idx])
        has_multiline = False
        
        # Parcourir toutes les lignes de données
        for row_idx in range(1, len(table_data)):
            row = table_data[row_idx]
            if col_idx < len(row):
                cell_text = str(row[col_idx])
                
                # Vérifier si la cellule contient plusieurs lignes
                if "\n" in cell_text:
                    has_multiline = True
                    lines = cell_text.split("\n")
                    # Prendre la ligne la plus longue
                    for line in lines:
                        max_chars = max(max_chars, len(line))
                else:
                    max_chars = max(max_chars, len(cell_text))
        
        # Facteur d'ajustement par type de colonne
        if col_idx == 0:  # Colonne Élève
            factor = 0.13  # cm par caractère
            min_width = 4 * cm
            max_width = 7 * cm
        elif col_idx == 2 and len(headers) > 3:  # Colonne des détails (Réductions/Dérogations/Pénalités)
            factor = 0.12 if not has_multiline else 0.14
            min_width = 3.5 * cm
            max_width = 6 * cm
        else:  # Colonnes de montants
            factor = 0.1
            min_width = 2.5 * cm
            max_width = 4 * cm
        
        # Calculer la largeur
        width = max_chars * factor * cm
        width = max(min_width, min(width, max_width))
        col_widths.append(width)
    
    return col_widths


def section_pdf(elements, titre, data, headers):
    """Ajoute une section au PDF avec largeurs dynamiques et gestion des sauts de ligne"""
    styles = getSampleStyleSheet()
    
    # Style pour le titre de section
    section_style = ParagraphStyle(
        'SectionTitle',
        parent=styles['Heading2'],
        fontSize=12,
        textColor=colors.HexColor('#2F528F'),
        spaceAfter=6,
        spaceBefore=10,
        alignment=TA_LEFT
    )
    
    # Titre de section
    elements.append(Paragraph(titre, section_style))
    
    if not data:
        elements.append(Paragraph("Aucun élève", styles['Italic']))
        elements.append(Spacer(1, 0.3*cm))
        return
    
    # Préparer les données du tableau
    table_data = [headers] + data
    
    # Calculer les largeurs dynamiques
    col_widths = calculate_dynamic_widths(table_data, headers)
    
    # Créer le tableau
    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    
    # Style de base du tableau
    style = [
        # En-tête
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2F528F')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('TOPPADDING', (0, 0), (-1, 0), 8),
        
        # Grille
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),  # Alignement en haut pour tous
        
        # Alignement des colonnes
        ('ALIGN', (0, 1), (0, -1), 'LEFT'),   # Élève à gauche
        ('ALIGN', (1, 1), (-1, -1), 'RIGHT'), # Montants à droite
    ]
    
    # Ajouter le style pour la colonne des détails (alignée à gauche)
    if len(headers) > 3:
        style.append(('ALIGN', (2, 1), (2, -1), 'LEFT'))
    
    table.setStyle(TableStyle(style))
    
    # Ajouter des couleurs alternées pour les lignes
    for i in range(1, len(table_data)):
        if i % 2 == 0:
            bg_color = colors.HexColor('#F9F9F9')
        else:
            bg_color = colors.HexColor('#FFFFFF')
        
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, i), (-1, i), bg_color),
        ]))
    
    elements.append(table)
    elements.append(Spacer(1, 0.5*cm))


def export_pdf_multi(tableau_complet, tableau_reduction, tableau_derogation, 
                    tableau_penalite, tableau_dette, headers_complet, 
                    headers_reduction, headers_derogation, headers_penalite, 
                    headers_dette, date_debut, date_fin, classe):
    
    buffer = BytesIO()
    
    # Créer le document PDF en format paysage avec marges confortables
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), 
                           leftMargin=1.2*cm, rightMargin=1.2*cm,
                           topMargin=1.5*cm, bottomMargin=1.2*cm)

    elements = []
    styles = getSampleStyleSheet()
    
    # Style pour le titre de section
    section_style = ParagraphStyle(
        'SectionTitle',
        parent=styles['Heading2'],
        fontSize=12,
        textColor=colors.HexColor('#2F528F'),
        spaceAfter=4,
        spaceBefore=8,
        alignment=TA_LEFT
    )
    
    # Style pour le nombre d'élèves
    count_style = ParagraphStyle(
        'CountStyle',
        parent=styles['Normal'],
        fontSize=9,
        textColor=colors.HexColor('#666666'),
        spaceAfter=4,
        alignment=TA_LEFT
    )
    
    # Construire le titre avec la période
    if date_debut and date_fin:
        titre = f"RAPPORT DES PAIEMENTS - Du {date_debut.strftime('%d/%m/%Y')} au {date_fin.strftime('%d/%m/%Y')}"
    elif date_debut:
        titre = f"RAPPORT DES PAIEMENTS - A partir du {date_debut.strftime('%d/%m/%Y')}"
    elif date_fin:
        titre = f"RAPPORT DES PAIEMENTS - Jusqu'au {date_fin.strftime('%d/%m/%Y')}"
    else:
        titre = "RAPPORT DES PAIEMENTS - Toute la période"
    
    # Extraire les IDs depuis l'objet classe
    campus_id = None
    annee_id = None
    cycle_id = None
    
    if classe:
        if hasattr(classe, 'id_campus'):
            campus_id = classe.id_campus.id_campus if hasattr(classe.id_campus, 'id_campus') else classe.id_campus
        if hasattr(classe, 'id_annee'):
            annee_id = classe.id_annee.id_annee if hasattr(classe.id_annee, 'id_annee') else classe.id_annee
        if hasattr(classe, 'cycle_id'):
            cycle_id = classe.cycle_id.id_cycle_actif if hasattr(classe.cycle_id, 'id_cycle_actif') else classe.cycle_id
    
    # Appeler la fonction header
    header_table = build_pdf_header(
        eleve=None,
        classe_obj=classe,
        id_campus=campus_id,
        id_cycle=cycle_id,
        id_annee=annee_id,
        titre=titre
    )
    
    elements.append(header_table)
    elements.append(Spacer(1, 0.8*cm))

    # ============================================
    # SECTIONS - Numérotation continue (1,2,3,4)
    # ============================================
    sections = []
    
    # N'ajouter que les sections qui ont des données
    if tableau_complet:
        sections.append(("1. PAIEMENTS COMPLETS", tableau_complet, headers_complet))
    if tableau_reduction:
        sections.append(("2. RÉDUCTIONS", tableau_reduction, headers_reduction))
    if tableau_derogation:
        sections.append(("3. DÉROGATIONS", tableau_derogation, headers_derogation))
    if tableau_penalite:
        sections.append(("4. PÉNALITÉS", tableau_penalite, headers_penalite))
    if tableau_dette:
        sections.append(("5. DETTES", tableau_dette, headers_dette))
    
    # Renumérotation continue
    for idx, (titre_section, data, headers) in enumerate(sections, 1):
        # Remplacer le premier chiffre par le nouvel index
        nouveau_titre = titre_section.replace(titre_section[0], str(idx), 1)
        
        # Titre de section
        elements.append(Paragraph(nouveau_titre, section_style))
        
        if data:
            # Afficher le tableau
            table_data = [headers] + data
            col_widths = calculate_dynamic_widths(table_data, headers)
            
            table = Table(table_data, colWidths=col_widths, repeatRows=1)
            
            style = [
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2F528F')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('TOPPADDING', (0, 0), (-1, 0), 8),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('ALIGN', (0, 1), (0, -1), 'LEFT'),
                ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),
            ]
            
            if len(headers) > 3:
                style.append(('ALIGN', (2, 1), (2, -1), 'LEFT'))
            
            table.setStyle(TableStyle(style))
            
            # Couleurs alternées
            for i in range(1, len(table_data)):
                if i % 2 == 0:
                    bg_color = colors.HexColor('#F9F9F9')
                else:
                    bg_color = colors.HexColor('#FFFFFF')
                
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, i), (-1, i), bg_color),
                ]))
            
            elements.append(table)
        
        elements.append(Spacer(1, 0.5*cm))
    
    # ============================================
    # RÉCAPITULATIF
    # ============================================
    elements.append(Paragraph("RÉCAPITULATIF", styles['Heading3']))
    elements.append(Spacer(1, 0.2*cm))
    
    total_eleves_classe = Eleve_inscription.objects.filter(id_classe=classe.id_classe_active).count()
    
    recap_data = [
        ["Catégorie", "Nombre d'élèves", "Pourcentage"],
        ["Complets", len(tableau_complet), f"{len(tableau_complet)/total_eleves_classe*100:.1f}%" if total_eleves_classe > 0 else "0%"],
        ["Réductions", len(tableau_reduction), f"{len(tableau_reduction)/total_eleves_classe*100:.1f}%" if total_eleves_classe > 0 else "0%"],
        ["Dérogations", len(tableau_derogation), f"{len(tableau_derogation)/total_eleves_classe*100:.1f}%" if total_eleves_classe > 0 else "0%"],
        ["Pénalités", len(tableau_penalite), f"{len(tableau_penalite)/total_eleves_classe*100:.1f}%" if total_eleves_classe > 0 else "0%"],
        ["Endettés", len(tableau_dette), f"{len(tableau_dette)/total_eleves_classe*100:.1f}%" if total_eleves_classe > 0 else "0%"],
        ["TOTAL", total_eleves_classe, "100%"]
    ]
    
    recap_table = Table(recap_data, colWidths=[5*cm, 3*cm, 3*cm])
    recap_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2F528F')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('BACKGROUND', (0, 1), (-1, -2), colors.HexColor('#F9F9F9')),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#E6E6E6')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    
    elements.append(recap_table)

    doc.build(elements)

    pdf = buffer.getvalue()
    buffer.close()

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="rapport_paiements.pdf"'
    response.write(pdf)

    return response